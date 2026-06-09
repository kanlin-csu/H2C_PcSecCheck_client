#!/usr/bin/env python3
"""
H2C_PcSecCheck_macos v2.0
macOS 主機資安健診工具 - 產生標準化 .h2cpc.zip 報告包

Copyright 2026 H2C工作室 甘霖老師
Licensed under the Apache License, Version 2.0
"""
from __future__ import annotations

import datetime
import hashlib
import html
import io
import json
import os
import platform
import plistlib
import re
import socket
import subprocess
import zipfile
from openpyxl import Workbook

TOOL_VERSION = "2.0.0"
SCHEMA_VERSION = "2"

SEVERITY_LABEL = {4: "Critical", 3: "High", 2: "Medium", 1: "Low", 0: "Info"}
SEVERITY_COLOR = {4: "#dc3545", 3: "#fd7e14", 2: "#ffc107", 1: "#0dcaf0", 0: "#6c757d"}
RISK_COLORS = {"critical": "#dc3545", "high": "#fd7e14", "medium": "#ffc107", "low": "#198754"}

_SUSPICIOUS_PATHS = re.compile(r"/(tmp|var/tmp|private/tmp|users/[^/]+/downloads)/", re.I)
_SUSPICIOUS_CMDLINE = (
    (r"curl.*\| *sh", 3),
    (r"wget.*\| *sh", 3),
    (r"bash -i", 4),
    (r"/dev/tcp/", 4),
    (r"nc -e", 4),
    (r"osascript.*do shell script", 3),
)
_PRIVATE_NETS = (
    re.compile(r"^10\."),
    re.compile(r"^172\.(1[6-9]|2\d|3[01])\."),
    re.compile(r"^192\.168\."),
    re.compile(r"^127\."),
    re.compile(r"^169\.254\."),
    re.compile(r"^::1$"),
    re.compile(r"^fe80:", re.I),
)
_MALICIOUS_PORTS = {4444, 5555, 6666, 1337, 31337, 9001}
_SENSITIVE_LISTEN = {22, 23, 445, 5900, 3283, 5000, 5432, 6379, 27017}


def _run(cmd: list[str], timeout: int = 15) -> str:
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return ((r.stdout or "") + (r.stderr or "")).replace("\r\n", "\n").replace("\r", "\n")
    except FileNotFoundError:
        return ""
    except Exception:
        return ""


def _read_text(path: str) -> str:
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            return f.read().replace("\r\n", "\n").replace("\r", "\n")
    except Exception:
        return ""


def _is_private_ip(ip: str) -> bool:
    ip = ip.strip("[]").split("%")[0]
    return any(p.match(ip) for p in _PRIVATE_NETS)


def get_local_ip_address() -> str:
    out = _run(["route", "-n", "get", "default"], timeout=5)
    iface = ""
    for line in out.splitlines():
        if "interface:" in line:
            iface = line.split(":", 1)[1].strip()
            break
    if iface:
        out = _run(["ipconfig", "getifaddr", iface], timeout=5).strip()
        if out and not out.startswith("127."):
            return out
    try:
        with socket.create_connection(("1.1.1.1", 80), timeout=3) as s:
            ip = s.getsockname()[0]
            if ip and not ip.startswith("127."):
                return ip
    except Exception:
        pass
    return "未知"


def get_system_info() -> list[dict]:
    data = [
        {"項目": "電腦名稱", "內容": platform.node()},
        {"項目": "區網 IP", "內容": get_local_ip_address()},
    ]
    for key, label in [
        ("ProductName", "產品名稱"),
        ("ProductVersion", "macOS 版本"),
        ("BuildVersion", "Build"),
    ]:
        val = _run(["sw_vers", f"-{key}"], timeout=5).strip()
        if val:
            data.append({"項目": label, "內容": val})
    data.append({"項目": "Kernel", "內容": platform.release()})
    data.append({"項目": "架構", "內容": platform.machine()})
    for name, cmd in [
        ("硬體型號", ["sysctl", "-n", "hw.model"]),
        ("CPU 核心數", ["sysctl", "-n", "hw.ncpu"]),
        ("記憶體 bytes", ["sysctl", "-n", "hw.memsize"]),
        ("運行時間", ["uptime"]),
    ]:
        out = _run(cmd, timeout=5).strip()
        if out:
            data.append({"項目": name, "內容": out})
    return data


def get_update_status() -> list[dict]:
    result = []
    out = _run(["softwareupdate", "-l"], timeout=60)
    if "No new software available" in out:
        result.append({"項目": "系統更新", "內容": "無可用更新"})
    elif out.strip():
        rows = [line.strip() for line in out.splitlines() if line.strip().startswith("*")]
        result.append({"項目": "可用系統更新數", "內容": str(len(rows))})
        for line in rows[:30]:
            result.append({"項目": "可用系統更新", "內容": line})
    else:
        result.append({"項目": "系統更新", "內容": "無法取得 softwareupdate 結果"})

    brew = _run(["brew", "outdated", "--formula"], timeout=30)
    if brew:
        rows = [x for x in brew.splitlines() if x.strip()]
        result.append({"項目": "Homebrew 可更新套件數", "內容": str(len(rows))})
        for line in rows[:30]:
            result.append({"項目": "Homebrew 可更新套件", "內容": line})
    return result


def get_endpoint_protection() -> list[dict]:
    result = []
    xprotect = _run(["/usr/libexec/PlistBuddy", "-c", "Print :CFBundleShortVersionString", "/Library/Apple/System/Library/CoreServices/XProtect.bundle/Contents/Info.plist"], timeout=5).strip()
    if not xprotect:
        xprotect = _run(["defaults", "read", "/Library/Apple/System/Library/CoreServices/XProtect.bundle/Contents/Info", "CFBundleShortVersionString"], timeout=5).strip()
    result.append({"項目": "XProtect", "狀態": "已偵測" if xprotect else "無法取得", "版本": xprotect})

    gatekeeper = _run(["spctl", "--status"], timeout=5).strip()
    result.append({"項目": "Gatekeeper", "狀態": gatekeeper or "無法取得", "版本": ""})

    mrt_paths = [
        "/Library/Apple/System/Library/CoreServices/MRT.app/Contents/Info.plist",
        "/System/Library/CoreServices/MRT.app/Contents/Info.plist",
    ]
    mrt_ver = ""
    for path in mrt_paths:
        if os.path.exists(path):
            mrt_ver = _run(["defaults", "read", path.replace(".plist", ""), "CFBundleShortVersionString"], timeout=5).strip()
            break
    result.append({"項目": "MRT", "狀態": "已偵測" if mrt_ver else "無法取得", "版本": mrt_ver})

    known = ("CrowdStrike", "SentinelOne", "Sophos", "TrendMicro", "Microsoft Defender", "JamfProtect")
    apps = []
    for base in ("/Applications", "/Library"):
        try:
            for name in os.listdir(base):
                low = name.lower()
                if any(k.lower().replace(" ", "") in low.replace(" ", "") for k in known):
                    apps.append(name)
        except Exception:
            pass
    result.append({"項目": "第三方端點防護", "狀態": ", ".join(sorted(set(apps))) if apps else "未偵測", "版本": ""})
    return result


def get_filevault_status() -> list[dict]:
    out = _run(["fdesetup", "status"], timeout=10).strip()
    return [{"項目": "FileVault", "狀態": out or "無法取得"}]


def get_sip_status() -> list[dict]:
    out = _run(["csrutil", "status"], timeout=10).strip()
    return [{"項目": "System Integrity Protection", "狀態": out or "無法取得"}]


def get_firewall_status() -> list[dict]:
    result = []
    fw = "/usr/libexec/ApplicationFirewall/socketfilterfw"
    state = _run([fw, "--getglobalstate"], timeout=10).strip()
    stealth = _run([fw, "--getstealthmode"], timeout=10).strip()
    block = _run([fw, "--getblockall"], timeout=10).strip()
    result.append({"項目": "Application Firewall", "狀態": state or "無法取得"})
    result.append({"項目": "Stealth Mode", "狀態": stealth or "無法取得"})
    result.append({"項目": "Block All", "狀態": block or "無法取得"})
    pf = _run(["pfctl", "-s", "info"], timeout=10)
    enabled = "Enabled" if "Status: Enabled" in pf else "Disabled" if "Status: Disabled" in pf else "未知"
    result.append({"項目": "PF Firewall", "狀態": enabled})
    return result


def get_local_users() -> list[dict]:
    result = []
    out = _run(["dscl", ".", "list", "/Users", "UniqueID"], timeout=15)
    admin_members = set(_run(["dscl", ".", "read", "/Groups/admin", "GroupMembership"], timeout=10).split()[1:])
    for line in out.splitlines():
        parts = line.split()
        if len(parts) < 2:
            continue
        name, uid_s = parts[0], parts[-1]
        if name.startswith("_"):
            account_type = "系統帳號"
        else:
            try:
                account_type = "一般使用者" if int(uid_s) >= 500 else "系統帳號"
            except Exception:
                account_type = "未知"
        shell = _run(["dscl", ".", "read", f"/Users/{name}", "UserShell"], timeout=5).replace("UserShell:", "").strip()
        home = _run(["dscl", ".", "read", f"/Users/{name}", "NFSHomeDirectory"], timeout=5).replace("NFSHomeDirectory:", "").strip()
        result.append({
            "帳號名稱": name,
            "UID": uid_s,
            "帳號類型": account_type,
            "是否管理員": "是" if name in admin_members else "否",
            "Shell": shell,
            "Home": home,
        })
    return result


def get_password_policy() -> list[dict]:
    result = []
    out = _run(["pwpolicy", "-getaccountpolicies"], timeout=10)
    if out.strip():
        for key in ("policyAttributeMinimumLength", "policyAttributeMaximumFailedAuthentications", "policyAttributeExpiresEveryNDays"):
            m = re.search(rf"<key>{key}</key>\s*<integer>(\d+)</integer>", out)
            result.append({"設定": key, "值": m.group(1) if m else "未設定"})
    else:
        result.append({"設定": "pwpolicy", "值": "無法取得"})
    return result


def get_ssh_security() -> list[dict]:
    cfg = {}
    text = _read_text("/etc/ssh/sshd_config")
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split(None, 1)
        if len(parts) == 2:
            cfg[parts[0].lower()] = parts[1]
    service = _run(["launchctl", "print", "system/com.openssh.sshd"], timeout=5)
    active = "已載入" if service else "未載入或未啟用"
    keys = {
        "permitrootlogin": "允許 root 登入",
        "passwordauthentication": "允許密碼登入",
        "permitemptypasswords": "允許空密碼",
        "pubkeyauthentication": "允許金鑰登入",
        "maxauthtries": "最大驗證嘗試次數",
    }
    result = [{"設定": "SSH 服務", "值": active, "狀態": "Info"}]
    for key, label in keys.items():
        result.append({"設定": label, "值": cfg.get(key, "未設定/預設"), "狀態": "待檢核"})
    return result


def get_launch_items() -> list[dict]:
    result = []
    paths = [
        ("/Library/LaunchDaemons", "LaunchDaemons"),
        ("/Library/LaunchAgents", "LaunchAgents"),
        ("/System/Library/LaunchDaemons", "System LaunchDaemons"),
        ("/System/Library/LaunchAgents", "System LaunchAgents"),
    ]
    for user_home in ("/Users/" + u for u in os.listdir("/Users") if not u.startswith(".")) if os.path.isdir("/Users") else []:
        paths.append((os.path.join(user_home, "Library/LaunchAgents"), "User LaunchAgents"))
    for folder, source in paths:
        if not os.path.isdir(folder):
            continue
        for fname in sorted(os.listdir(folder)):
            if not fname.endswith(".plist"):
                continue
            fpath = os.path.join(folder, fname)
            label = fname
            program = ""
            try:
                with open(fpath, "rb") as f:
                    pl = plistlib.load(f)
                label = pl.get("Label", fname)
                program = pl.get("Program", "") or " ".join(str(x) for x in pl.get("ProgramArguments", []))
            except Exception:
                pass
            result.append({"來源": source, "名稱": label, "狀態": "啟用/存在", "命令/路徑": program or fpath})
    return result


def get_processes() -> list[dict]:
    out = _run(["ps", "-axo", "pid,ppid,comm,args"], timeout=20)
    result = []
    for line in out.splitlines()[1:]:
        parts = line.strip().split(None, 3)
        if len(parts) < 3:
            continue
        pid, ppid, comm = parts[:3]
        args = parts[3] if len(parts) > 3 else ""
        result.append({
            "程序名稱": os.path.basename(comm),
            "PID": pid,
            "父PID": ppid,
            "執行路徑": comm,
            "命令列": args[:600],
        })
    return result


def _empty_threat_fields() -> dict:
    return {"vt_status": "not_queried", "threat_source": "", "first_seen": "", "last_seen": ""}


def _package_owner(path: str) -> str:
    if not path or not os.path.exists(path):
        return ""
    out = _run(["pkgutil", "--file-info", path], timeout=5)
    for line in out.splitlines():
        if line.startswith("pkgid:"):
            return line.split(":", 1)[1].strip()
    return ""


def get_process_hashes() -> list[dict]:
    result = []
    seen = set()
    for proc in get_processes():
        path = proc.get("執行路徑", "")
        if not path or path in seen or not os.path.isfile(path):
            continue
        seen.add(path)
        row = {
            "程序名稱": proc.get("程序名稱", ""),
            "PID": proc.get("PID", ""),
            "執行路徑": path,
            "SHA256": "",
            "hash_status": "not_calculated",
            "套件來源": "",
            "是否已刪除": "否",
            **_empty_threat_fields(),
        }
        try:
            sha = hashlib.sha256()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    sha.update(chunk)
            row["SHA256"] = sha.hexdigest().upper()
            row["hash_status"] = "ok"
            row["套件來源"] = _package_owner(path)
        except PermissionError:
            row["hash_status"] = "permission_denied"
        except Exception:
            row["hash_status"] = "failed"
        result.append(row)
    return result


def get_netstat() -> list[dict]:
    result = []
    out = _run(["lsof", "-nP", "-iTCP", "-iUDP"], timeout=20)
    for line in out.splitlines()[1:]:
        parts = line.split()
        if len(parts) < 9:
            continue
        proc, pid = parts[0], parts[1]
        proto = parts[7] if len(parts) > 7 else ""
        name = " ".join(parts[8:])
        state = ""
        m = re.search(r"\(([^)]+)\)", name)
        if m:
            state = m.group(1).upper()
            name = name.replace(m.group(0), "").strip()
        local, remote = name, ""
        if "->" in name:
            local, remote = [x.strip() for x in name.split("->", 1)]
        result.append({"協定": proto, "本地位址": local, "遠端位址": remote, "狀態": state, "PID": pid, "程序名稱": proc})
    if result:
        return result
    out = _run(["netstat", "-anv"], timeout=20)
    for line in out.splitlines():
        if not line.startswith(("tcp", "udp")):
            continue
        parts = line.split()
        if len(parts) >= 4:
            result.append({"協定": parts[0].upper(), "本地位址": parts[3], "遠端位址": parts[4] if len(parts) > 4 else "", "狀態": parts[5] if len(parts) > 5 else "", "PID": "", "程序名稱": ""})
    return result


def get_network_shares() -> list[dict]:
    result = []
    out = _run(["sharing", "-l"], timeout=15)
    current = {}
    for line in out.splitlines():
        line = line.strip()
        if line.startswith("name:"):
            if current:
                result.append(current)
            current = {"共用名稱": line.split(":", 1)[1].strip(), "路徑": "", "說明": "", "類型": "macOS sharing"}
        elif line.startswith("path:") and current:
            current["路徑"] = line.split(":", 1)[1].strip()
    if current:
        result.append(current)
    exports = _read_text("/etc/exports")
    for line in exports.splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            parts = line.split(None, 1)
            result.append({"共用名稱": os.path.basename(parts[0]) or parts[0], "路徑": parts[0], "說明": parts[1] if len(parts) > 1 else "", "類型": "NFS"})
    return result


def get_audit_policy() -> list[dict]:
    result = []
    audit_control = _read_text("/etc/security/audit_control")
    if audit_control:
        for line in audit_control.splitlines():
            if line.strip() and not line.startswith("#"):
                k, _, v = line.partition(":")
                result.append({"稽核類別": k.strip(), "稽核設定": v.strip()})
    else:
        result.append({"稽核類別": "audit_control", "稽核設定": "無法讀取"})
    auditd = _run(["launchctl", "print", "system/com.apple.auditd"], timeout=5)
    result.append({"稽核類別": "auditd 服務", "稽核設定": "已載入" if auditd else "未載入或無法取得"})
    return result


def get_kernel_hardening() -> list[dict]:
    checks = {
        "kern.securelevel": "securelevel",
        "net.inet.ip.forwarding": "IPv4 轉送",
        "net.inet6.ip6.forwarding": "IPv6 轉送",
        "net.inet.tcp.blackhole": "TCP blackhole",
        "net.inet.udp.blackhole": "UDP blackhole",
    }
    return [{"項目": label, "sysctl": key, "值": _run(["sysctl", "-n", key], timeout=5).strip() or "無法取得"} for key, label in checks.items()]


def get_hosts_file() -> list[dict]:
    result = []
    std_hosts = {"localhost", "broadcasthost"}
    std_ips = {"127.0.0.1", "::1", "255.255.255.255"}
    try:
        with open("/etc/hosts", encoding="utf-8", errors="replace") as f:
            for lineno, line in enumerate(f, 1):
                stripped = line.strip()
                if not stripped or stripped.startswith("#"):
                    continue
                parts = stripped.split()
                if len(parts) >= 2:
                    ip, host = parts[0], parts[1]
                    result.append({"行號": lineno, "IP": ip, "主機名稱": host, "完整內容": stripped, "是否標準": "是" if ip in std_ips and host in std_hosts else "否"})
    except Exception as e:
        result.append({"行號": 0, "IP": "", "主機名稱": "", "完整內容": f"無法讀取: {e}", "是否標準": "否"})
    return result


def _finding(category: str, severity: int, title: str, detail: str, raw_data: dict | None = None) -> dict:
    return {"category": category, "severity": severity, "title": title, "detail": detail, "raw_data": raw_data or {}, "status": "open"}


def _analyze_updates(rows: list[dict]) -> list[dict]:
    findings = []
    data = {r.get("項目", ""): r.get("內容", "") for r in rows}
    for key in ("可用系統更新數", "Homebrew 可更新套件數"):
        try:
            count = int(data.get(key, "0"))
            if count > 0:
                findings.append(_finding("macos_patch_management", 2 if count < 20 else 3, f"{key}: {count}", "請確認是否包含安全性修補並安排更新。", {key: count}))
        except Exception:
            pass
    return findings


def _analyze_endpoint(rows: list[dict]) -> list[dict]:
    findings = []
    data = {r.get("項目", ""): r.get("狀態", "") for r in rows}
    if "disabled" in data.get("Gatekeeper", "").lower():
        findings.append(_finding("macos_endpoint_protection", 3, "Gatekeeper 已停用", "建議啟用 Gatekeeper，降低未簽章程式執行風險。", data))
    if data.get("第三方端點防護") == "未偵測":
        findings.append(_finding("macos_endpoint_protection", 1, "未偵測到第三方端點防護", "若機關政策要求 EDR/AV，請確認是否另有集中式防護。", data))
    return findings


def _analyze_filevault(rows: list[dict]) -> list[dict]:
    state = rows[0].get("狀態", "") if rows else ""
    if "FileVault is On" not in state:
        return [_finding("macos_disk_encryption", 3, "FileVault 未啟用", f"目前狀態：{state}", rows[0] if rows else {})]
    return []


def _analyze_sip(rows: list[dict]) -> list[dict]:
    state = rows[0].get("狀態", "") if rows else ""
    if "enabled" not in state.lower():
        return [_finding("macos_system_integrity", 4, "SIP 未啟用", f"目前狀態：{state}", rows[0] if rows else {})]
    return []


def _analyze_firewall(rows: list[dict]) -> list[dict]:
    findings = []
    data = {r.get("項目", ""): r.get("狀態", "") for r in rows}
    if "enabled" not in data.get("Application Firewall", "").lower():
        findings.append(_finding("macos_firewall", 3, "Application Firewall 未啟用", "建議啟用 macOS Application Firewall。", data))
    if data.get("PF Firewall") == "Disabled":
        findings.append(_finding("macos_firewall", 1, "PF Firewall 未啟用", "若有主機型防火牆政策，請確認 PF 或其他防火牆設定。", data))
    return findings


def _analyze_users(users: list[dict]) -> list[dict]:
    findings = []
    admins = [u for u in users if u.get("是否管理員") == "是"]
    for u in admins:
        findings.append(_finding("macos_privilege_management", 1, f"管理員帳號：{u.get('帳號名稱')}", "管理員帳號應定期盤點並確認授權紀錄。", u))
    for u in users:
        if u.get("帳號類型") == "一般使用者":
            findings.append(_finding("macos_account_inventory", 0, f"非系統帳號：{u.get('帳號名稱')}", "請確認帳號用途、人員歸屬與離職停用流程。", u))
    return findings


def _analyze_password_policy(rows: list[dict]) -> list[dict]:
    findings = []
    data = {r.get("設定", ""): r.get("值", "") for r in rows}
    try:
        min_len = int(data.get("policyAttributeMinimumLength", "0"))
        if min_len < 8:
            findings.append(_finding("macos_password_policy", 2, "密碼最短長度不足", f"目前最短長度：{min_len}，建議至少 8。", data))
    except Exception:
        findings.append(_finding("macos_password_policy", 1, "無法確認密碼最短長度", "請人工確認 pwpolicy 或 MDM 密碼政策。", data))
    return findings


def _analyze_ssh(rows: list[dict]) -> list[dict]:
    findings = []
    data = {r.get("設定", ""): str(r.get("值", "")).lower() for r in rows}
    if data.get("允許 root 登入") == "yes":
        findings.append(_finding("macos_ssh_security", 3, "SSH 允許 root 登入", "建議停用 root 直接登入。", data))
    if data.get("允許密碼登入") == "yes":
        findings.append(_finding("macos_ssh_security", 2, "SSH 允許密碼登入", "建議使用金鑰登入並搭配來源限制。", data))
    if data.get("允許空密碼") == "yes":
        findings.append(_finding("macos_ssh_security", 4, "SSH 允許空密碼", "應立即停用 PermitEmptyPasswords。", data))
    return findings


def _analyze_processes(processes: list[dict]) -> list[dict]:
    findings = []
    for proc in processes:
        path = (proc.get("執行路徑") or "").lower()
        cmd = (proc.get("命令列") or "").lower()
        if _SUSPICIOUS_PATHS.search(path):
            findings.append(_finding("macos_suspicious_process", 3, f"程序執行於可疑路徑：{proc.get('程序名稱')}", f"執行路徑：{proc.get('執行路徑')}", proc))
        for pattern, sev in _SUSPICIOUS_CMDLINE:
            if re.search(pattern, cmd, re.I):
                findings.append(_finding("macos_suspicious_process", sev, f"程序命令列含可疑特徵：{proc.get('程序名稱')}", f"符合規則：{pattern}\n{proc.get('命令列', '')[:300]}", proc))
                break
    return findings


def _parse_port(addr: str) -> int:
    try:
        return int(addr.rsplit(":", 1)[1])
    except Exception:
        return 0


def _parse_ip(addr: str) -> str:
    if "->" in addr:
        addr = addr.split("->", 1)[0].strip()
    return addr.rsplit(":", 1)[0].strip("[]")


def _analyze_netstat(rows: list[dict]) -> list[dict]:
    findings = []
    for row in rows:
        local = row.get("本地位址", "")
        remote = row.get("遠端位址", "")
        state = row.get("狀態", "")
        local_port = _parse_port(local)
        remote_port = _parse_port(remote)
        local_ip = _parse_ip(local)
        remote_ip = _parse_ip(remote)
        if state in ("LISTEN", "LISTENING") and local_port in _SENSITIVE_LISTEN and local_ip in ("*", "", "0.0.0.0", "::"):
            findings.append(_finding("macos_service_exposure", 2, f"敏感服務對外監聽：{local_port}", f"{row.get('程序名稱')} 監聽 {local}", row))
        if state in ("ESTABLISHED", "ESTAB") and remote_ip and not _is_private_ip(remote_ip):
            if remote_port in _MALICIOUS_PORTS:
                findings.append(_finding("macos_suspicious_connection", 3, f"連線至可疑 Port：{remote_port}", f"{row.get('程序名稱')} -> {remote}", row))
    return findings


def _analyze_launch_items(rows: list[dict]) -> list[dict]:
    findings = []
    for item in rows:
        cmd = (item.get("命令/路徑") or "").lower()
        if _SUSPICIOUS_PATHS.search(cmd):
            findings.append(_finding("macos_persistence", 3, f"啟動項目指向可疑路徑：{item.get('名稱')}", f"{item.get('來源')} -> {item.get('命令/路徑')}", item))
    return findings


def _analyze_audit(rows: list[dict]) -> list[dict]:
    data = {r.get("稽核類別", ""): r.get("稽核設定", "") for r in rows}
    findings = []
    if data.get("auditd 服務") != "已載入":
        findings.append(_finding("macos_audit_logging", 2, "auditd 未載入或無法確認", "請確認 macOS auditd 與 audit_control 設定。", data))
    flags = data.get("flags", "")
    if not flags:
        findings.append(_finding("macos_audit_logging", 1, "audit flags 未設定", "請確認 /etc/security/audit_control 是否符合稽核政策。", data))
    return findings


def _analyze_shares(rows: list[dict]) -> list[dict]:
    return [_finding("macos_network_share", 2, f"偵測到網路分享：{r.get('共用名稱')}", f"路徑：{r.get('路徑')}，請確認權限與業務必要性。", r) for r in rows]


def _analyze_hosts(rows: list[dict]) -> list[dict]:
    non_std = [h for h in rows if h.get("是否標準") == "否" and h.get("IP")]
    if non_std:
        return [_finding("macos_hosts_file", 2, f"hosts 檔案含 {len(non_std)} 筆非標準記錄", "請確認是否為合法內部解析或重導向。\n" + "\n".join(f"{h['IP']} {h['主機名稱']}" for h in non_std[:20]), {"非標準記錄": non_std})]
    return []


def analyze_findings(report_data: dict) -> list[dict]:
    findings = []
    findings += _analyze_updates(report_data.get("更新狀態", []))
    findings += _analyze_endpoint(report_data.get("端點防護", []))
    findings += _analyze_filevault(report_data.get("FileVault", []))
    findings += _analyze_sip(report_data.get("SIP狀態", []))
    findings += _analyze_firewall(report_data.get("防火牆狀態", []))
    findings += _analyze_users(report_data.get("使用者帳號", []))
    findings += _analyze_password_policy(report_data.get("密碼原則", []))
    findings += _analyze_ssh(report_data.get("SSH安全設定", []))
    findings += _analyze_processes(report_data.get("執行中程序", []))
    findings += _analyze_netstat(report_data.get("網路連線", []))
    findings += _analyze_launch_items(report_data.get("啟動項目", []))
    findings += _analyze_audit(report_data.get("稽核原則", []))
    findings += _analyze_shares(report_data.get("網路分享", []))
    findings += _analyze_hosts(report_data.get("Hosts檔案", []))
    return findings


def calculate_risk(findings: list[dict]) -> tuple[int, str]:
    sev_score = {4: 40, 3: 20, 2: 10, 1: 3, 0: 0}
    score = min(100, sum(sev_score.get(f.get("severity", 0), 0) for f in findings))
    if score >= 80:
        return score, "critical"
    if score >= 50:
        return score, "high"
    if score >= 20:
        return score, "medium"
    return score, "low"


def build_compliance_summary(report_data: dict, findings: list[dict]) -> list[dict]:
    def row(item: str, value: str, status: str, suggestion: str) -> dict:
        return {"檢核項目": item, "目前狀態": value, "判定": status, "建議": suggestion}

    fw = {x.get("項目", ""): x.get("狀態", "") for x in report_data.get("防火牆狀態", [])}
    ep = {x.get("項目", ""): x.get("狀態", "") for x in report_data.get("端點防護", [])}
    fv = report_data.get("FileVault", [{}])[0].get("狀態", "")
    sip = report_data.get("SIP狀態", [{}])[0].get("狀態", "")
    users = report_data.get("使用者帳號", [])
    admins = [u for u in users if u.get("是否管理員") == "是"]
    updates = {x.get("項目", ""): x.get("內容", "") for x in report_data.get("更新狀態", [])}
    ssh = {x.get("設定", ""): x.get("值", "") for x in report_data.get("SSH安全設定", [])}
    audit = {x.get("稽核類別", ""): x.get("稽核設定", "") for x in report_data.get("稽核原則", [])}
    listen = [n for n in report_data.get("網路連線", []) if n.get("狀態") in ("LISTEN", "LISTENING")]
    shares = report_data.get("網路分享", [])

    return [
        row("系統更新", updates.get("可用系統更新數", updates.get("系統更新", "未知")), "WARN" if updates.get("可用系統更新數") not in ("0", None, "") else "PASS", "請確認是否包含安全性修補並保留更新紀錄。"),
        row("FileVault 磁碟加密", fv, "PASS" if "On" in fv else "FAIL", "建議啟用 FileVault。"),
        row("SIP 系統完整性保護", sip, "PASS" if "enabled" in sip.lower() else "FAIL", "建議保持 SIP 啟用。"),
        row("Gatekeeper", ep.get("Gatekeeper", "未知"), "FAIL" if "disabled" in ep.get("Gatekeeper", "").lower() else "PASS", "建議啟用 Gatekeeper。"),
        row("Application Firewall", fw.get("Application Firewall", "未知"), "PASS" if "enabled" in fw.get("Application Firewall", "").lower() else "FAIL", "建議啟用 macOS 防火牆。"),
        row("管理員帳號", f"{len(admins)} 個", "WARN" if len(admins) > 1 else "PASS", "管理員帳號需定期盤點與授權。"),
        row("SSH 安全", f"RootLogin={ssh.get('允許 root 登入','未偵測')}, PasswordAuth={ssh.get('允許密碼登入','未偵測')}", "WARN" if str(ssh.get("允許 root 登入", "")).lower() == "yes" or str(ssh.get("允許密碼登入", "")).lower() == "yes" else "PASS", "建議停用 root 直接登入並使用金鑰。"),
        row("稽核記錄", audit.get("auditd 服務", "未知"), "PASS" if audit.get("auditd 服務") == "已載入" else "WARN", "請確認 audit_control 是否符合政策。"),
        row("對外監聽服務", f"{len(listen)} 個", "INFO" if listen else "PASS", "請確認每個 LISTEN 服務的業務必要性。"),
        row("網路分享 / 對外目錄", f"{len(shares)} 筆", "WARN" if shares else "PASS", "請確認分享權限與最小權限原則。"),
    ]


_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Noto Sans TC', Arial, sans-serif; font-size: 13px; background: #f0f2f5; color: #212529; }
.topbar { background: #212529; color: #fff; padding: 10px 20px; position: sticky; top: 0; z-index: 100; display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
.topbar .brand { font-size: 15px; font-weight: 700; margin-right: 8px; }
.topbar a { color: #adb5bd; text-decoration: none; padding: 4px 8px; border-radius: 4px; font-size: 12px; }
.topbar a:hover { background: #343a40; color: #fff; }
.container { max-width: 1440px; margin: 0 auto; padding: 20px; }
h1 { font-size: 20px; margin: 16px 0 4px; }
.meta-line { color: #6c757d; margin-bottom: 20px; font-size: 12px; }
.card { background: #fff; border: 1px solid #dee2e6; border-radius: 6px; margin-bottom: 22px; }
.card-header { padding: 9px 16px; background: #343a40; color: #fff; border-radius: 6px 6px 0 0; font-weight: 600; display: flex; justify-content: space-between; align-items: center; }
.card-body { padding: 14px; }
.summary-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(120px, 1fr)); gap: 10px; margin-bottom: 18px; }
.s-item { background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px; padding: 10px; text-align: center; }
.s-item .val { font-size: 26px; font-weight: 700; }
.s-item .lbl { font-size: 11px; color: #6c757d; margin-top: 3px; }
.badge { display: inline-block; padding: 3px 10px; border-radius: 12px; color: #fff; font-size: 11px; font-weight: 600; }
.status { display:inline-block; min-width:56px; text-align:center; padding:3px 8px; border-radius:10px; color:#fff; font-weight:700; font-size:11px; }
.st-PASS { background:#198754; }
.st-WARN { background:#ffc107; color:#212529; }
.st-FAIL { background:#dc3545; }
.st-INFO { background:#0dcaf0; color:#212529; }
.search { margin-bottom: 8px; }
.search input { padding: 5px 10px; border: 1px solid #ced4da; border-radius: 4px; font-size: 12px; width: 100%; max-width: 300px; }
.tbl-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 12px; }
th { background: #495057; color: #fff; padding: 6px 8px; text-align: left; cursor: pointer; white-space: nowrap; user-select: none; }
td { padding: 5px 8px; border-bottom: 1px solid #e9ecef; word-break: break-word; max-width: 420px; vertical-align: top; }
tr:hover td { background: #f8f9fa; }
"""

_JS = """
function sortT(id, col) {
  var t = document.getElementById(id); if (!t) return;
  var dir = t.dataset.dir === "asc" ? "desc" : "asc"; t.dataset.dir = dir;
  var rows = Array.from(t.querySelectorAll("tbody tr"));
  rows.sort(function(a, b) {
    var x = (a.cells[col] || {}).innerText || "";
    var y = (b.cells[col] || {}).innerText || "";
    var c = x.localeCompare(y, "zh-Hant", {numeric: true, sensitivity: "base"});
    return dir === "asc" ? c : -c;
  });
  var tb = t.querySelector("tbody"); rows.forEach(function(r) { tb.appendChild(r); });
}
function filterT(inp, id) {
  var f = inp.value.toLowerCase();
  document.getElementById(id).querySelectorAll("tbody tr").forEach(function(r) {
    r.style.display = r.innerText.toLowerCase().includes(f) ? "" : "none";
  });
}
"""


def _section_table(data: list, tid: str) -> str:
    if not data:
        return '<p style="color:#6c757d;padding:8px">無資料</p>'
    if not (isinstance(data, list) and isinstance(data[0], dict)):
        return "".join(f"<p>{html.escape(str(item))}</p>" for item in data)
    headers = list(data[0].keys())
    h = [f'<div class="search"><input placeholder="搜尋..." oninput="filterT(this,\'{tid}\')"></div>', f'<div class="tbl-wrap"><table id="{tid}"><thead><tr>']
    for i, hdr in enumerate(headers):
        h.append(f'<th onclick="sortT(\'{tid}\',{i})">{hdr} ↕</th>')
    h.append("</tr></thead><tbody>")
    for item in data:
        h.append("<tr>")
        for hdr in headers:
            h.append(f"<td>{html.escape(str(item.get(hdr, '')))}</td>")
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "\n".join(h)


def _focus_table(summary: list[dict]) -> str:
    headers = ["檢核項目", "目前狀態", "判定", "建議"]
    h = ['<div class="tbl-wrap"><table id="focus_t"><thead><tr>']
    for i, hdr in enumerate(headers):
        h.append(f'<th onclick="sortT(\'focus_t\',{i})">{hdr} ↕</th>')
    h.append("</tr></thead><tbody>")
    for item in summary:
        status = str(item.get("判定", "INFO")).upper()
        h.append("<tr>")
        h.append(f"<td>{html.escape(str(item.get('檢核項目', '')))}</td>")
        h.append(f"<td>{html.escape(str(item.get('目前狀態', '')))}</td>")
        h.append(f'<td><span class="status st-{status}">{status}</span></td>')
        h.append(f"<td>{html.escape(str(item.get('建議', '')))}</td>")
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "\n".join(h)


def generate_html(computer_name: str, local_ip: str, report_data: dict, findings: list, risk_score: int, risk_level: str) -> str:
    sc = {s: sum(1 for f in findings if f.get("severity") == s) for s in (4, 3, 2, 1, 0)}
    risk_color = RISK_COLORS.get(risk_level, "#6c757d")
    focus = build_compliance_summary(report_data, findings)

    def findings_rows() -> str:
        rows = []
        for f in sorted(findings, key=lambda x: -x.get("severity", 0)):
            sev = f.get("severity", 0)
            badge = f'<span class="badge" style="background:{SEVERITY_COLOR.get(sev, "#6c757d")}">{SEVERITY_LABEL.get(sev, "?")}</span>'
            detail = html.escape(f.get("detail") or "").replace("\n", "<br>")
            rows.append(f"<tr><td>{badge}</td><td>{html.escape(f.get('category',''))}</td><td>{html.escape(f.get('title',''))}</td><td>{detail}</td></tr>")
        return "\n".join(rows) if rows else '<tr><td colspan="4" style="color:#6c757d;text-align:center">無風險項目</td></tr>'

    sections = [
        ("系統資訊", "sys_t", report_data.get("系統資訊", [])),
        ("更新狀態", "upd_t", report_data.get("更新狀態", [])),
        ("端點防護", "ep_t", report_data.get("端點防護", [])),
        ("FileVault", "fv_t", report_data.get("FileVault", [])),
        ("SIP 狀態", "sip_t", report_data.get("SIP狀態", [])),
        ("防火牆狀態", "fw_t", report_data.get("防火牆狀態", [])),
        ("帳號與權限", "usr_t", report_data.get("使用者帳號", [])),
        ("密碼原則", "pol_t", report_data.get("密碼原則", [])),
        ("SSH 安全設定", "ssh_t", report_data.get("SSH安全設定", [])),
        ("啟動項目", "launch_t", report_data.get("啟動項目", [])),
        ("稽核原則", "audit_t", report_data.get("稽核原則", [])),
        ("Kernel Hardening", "kernel_t", report_data.get("KernelHardening", [])),
        ("網路分享 / 對外目錄", "share_t", report_data.get("網路分享", [])),
        ("網路連線", "conn_t", report_data.get("網路連線", [])),
        ("Hosts 檔案", "hosts_t", report_data.get("Hosts檔案", [])),
        ("程序情資素材 / Hash 清單", "hash_t", report_data.get("程序雜湊", [])),
        ("執行中程序", "proc_t", report_data.get("執行中程序", [])),
    ]
    nav = " ".join(f'<a href="#{aid}">{label}</a>' for label, aid in [("健診重點", "focus_sec"), ("風險摘要", "findings_sec")] + [(t, f"sec_{i}") for t, i, _ in sections])
    cards = "".join(f'<div class="card" id="sec_{tid}"><div class="card-header">{title}</div><div class="card-body">{_section_table(data, tid)}</div></div>' for title, tid, data in sections)
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"""<!DOCTYPE html>
<html lang="zh-Hant"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>H2C macOS 健診 - {computer_name} ({local_ip})</title><style>{_CSS}</style></head>
<body><div class="topbar"><span class="brand">H2C macOS 健診</span>{nav}</div>
<div class="container"><h1>macOS 資安健診報告</h1>
<p class="meta-line">主機：<strong>{computer_name}</strong> &nbsp;|&nbsp; IP：<strong>{local_ip}</strong> &nbsp;|&nbsp; 產生時間：{now_str}</p>
<div class="card" id="focus_sec"><div class="card-header">行政院資安健診重點摘要</div><div class="card-body">{_focus_table(focus)}</div></div>
<div class="card" id="findings_sec"><div class="card-header">風險摘要 <span class="badge" style="background:{risk_color};font-size:13px;padding:4px 14px">{risk_level.upper()} {risk_score}/100</span></div>
<div class="card-body"><div class="summary-grid">
<div class="s-item"><div class="val" style="color:#dc3545">{sc[4]}</div><div class="lbl">Critical</div></div>
<div class="s-item"><div class="val" style="color:#fd7e14">{sc[3]}</div><div class="lbl">High</div></div>
<div class="s-item"><div class="val" style="color:#ffc107">{sc[2]}</div><div class="lbl">Medium</div></div>
<div class="s-item"><div class="val" style="color:#0dcaf0">{sc[1]}</div><div class="lbl">Low</div></div>
<div class="s-item"><div class="val" style="color:#6c757d">{sc[0]}</div><div class="lbl">Info</div></div>
</div><div class="tbl-wrap"><table id="findings_t"><thead><tr><th>嚴重度</th><th>類別</th><th>標題</th><th>說明</th></tr></thead><tbody>{findings_rows()}</tbody></table></div></div></div>
{cards}</div><script>{_JS}</script></body></html>"""


def _safe_cell(val):
    s = str(val) if not isinstance(val, (int, float, type(None))) else val
    if isinstance(s, str) and s and s[0] in ("=", "+", "-", "@", "|", "%"):
        return "'" + s
    return val


def _write_sheet(ws, data: list, field_order: list | None = None) -> None:
    if not data:
        ws.append(["無資料"])
        return
    if field_order is None:
        field_order = list(data[0].keys()) if data else []
    ws.append(field_order)
    for item in data:
        ws.append([_safe_cell(item.get(field, "")) for field in field_order])


def generate_xlsx(report_data: dict, findings: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "健診重點"
    _write_sheet(ws, build_compliance_summary(report_data, findings), ["檢核項目", "目前狀態", "判定", "建議"])
    ws = wb.create_sheet("風險摘要")
    ws.append(["類別", "嚴重度", "標題", "說明", "狀態"])
    for f in sorted(findings, key=lambda x: -x.get("severity", 0)):
        ws.append([_safe_cell(f.get("category", "")), SEVERITY_LABEL.get(f.get("severity", 0), ""), _safe_cell(f.get("title", "")), _safe_cell(f.get("detail", "")), f.get("status", "open")])
    for sheet_name, key in [
        ("系統資訊", "系統資訊"), ("更新狀態", "更新狀態"), ("端點防護", "端點防護"), ("FileVault", "FileVault"),
        ("SIP狀態", "SIP狀態"), ("防火牆狀態", "防火牆狀態"), ("使用者帳號", "使用者帳號"), ("密碼原則", "密碼原則"),
        ("SSH安全設定", "SSH安全設定"), ("啟動項目", "啟動項目"), ("稽核原則", "稽核原則"), ("KernelHardening", "KernelHardening"),
        ("網路分享", "網路分享"), ("網路連線", "網路連線"), ("Hosts檔案", "Hosts檔案"), ("程序情資素材", "程序雜湊"), ("執行中程序", "執行中程序"),
    ]:
        _write_sheet(wb.create_sheet(sheet_name), report_data.get(key, []))
    return wb


def package_zip(base_name: str, report_data: dict, findings: list, risk_score: int, risk_level: str, computer_name: str, local_ip: str) -> str:
    report_json_bytes = json.dumps(report_data, ensure_ascii=False, indent=2).encode("utf-8")
    findings_with_id = [dict(f, id=f"macos:{f['category']}:{i}") for i, f in enumerate(findings)]
    findings_json_bytes = json.dumps(findings_with_id, ensure_ascii=False, indent=2).encode("utf-8")
    html_bytes = generate_html(computer_name, local_ip, report_data, findings, risk_score, risk_level).encode("utf-8")
    buf = io.BytesIO()
    generate_xlsx(report_data, findings).save(buf)
    xlsx_bytes = buf.getvalue()
    meta = {
        "tool": "H2C_PcSecCheck_macos",
        "version": TOOL_VERSION,
        "schema_ver": SCHEMA_VERSION,
        "computer_name": computer_name,
        "local_ip": local_ip,
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "risk_score": risk_score,
        "risk_level": risk_level,
        "checksums": {
            f"{base_name}.report.json": hashlib.sha256(report_json_bytes).hexdigest(),
            f"{base_name}.findings.json": hashlib.sha256(findings_json_bytes).hexdigest(),
        },
    }
    zip_filename = f"{base_name}.h2cpc.zip"
    with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("meta.json", json.dumps(meta, ensure_ascii=False, indent=2).encode("utf-8"))
        zf.writestr(f"{base_name}.report.json", report_json_bytes)
        zf.writestr(f"{base_name}.findings.json", findings_json_bytes)
        zf.writestr(f"{base_name}.report.html", html_bytes)
        zf.writestr(f"{base_name}.report.xlsx", xlsx_bytes)
    return zip_filename


def _print(msg: str) -> None:
    print(msg, flush=True)


def main() -> None:
    if os.geteuid() != 0:
        _print("[警告] 建議以 sudo 執行，否則部分資訊可能無法完整收集。")
        _print("")
    computer_name = platform.node()
    local_ip = get_local_ip_address()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"{computer_name}_{local_ip}_{ts}"
    _print("=" * 56)
    _print("  H2C macOS 資安健診工具  v2.0")
    _print(f"  主機：{computer_name}   IP：{local_ip}")
    _print("=" * 56)
    steps = [
        ("[1/17] 取得系統資訊...", get_system_info, "系統資訊"),
        ("[2/17] 取得更新狀態...", get_update_status, "更新狀態"),
        ("[3/17] 取得端點防護...", get_endpoint_protection, "端點防護"),
        ("[4/17] 取得 FileVault...", get_filevault_status, "FileVault"),
        ("[5/17] 取得 SIP 狀態...", get_sip_status, "SIP狀態"),
        ("[6/17] 取得防火牆狀態...", get_firewall_status, "防火牆狀態"),
        ("[7/17] 取得使用者帳號...", get_local_users, "使用者帳號"),
        ("[8/17] 取得密碼原則...", get_password_policy, "密碼原則"),
        ("[9/17] 取得 SSH 安全設定...", get_ssh_security, "SSH安全設定"),
        ("[10/17] 取得啟動項目...", get_launch_items, "啟動項目"),
        ("[11/17] 取得執行中程序...", get_processes, "執行中程序"),
        ("[12/17] 收集程序情資素材 / SHA256...", get_process_hashes, "程序雜湊"),
        ("[13/17] 取得網路連線...", get_netstat, "網路連線"),
        ("[14/17] 取得網路分享...", get_network_shares, "網路分享"),
        ("[15/17] 取得稽核原則...", get_audit_policy, "稽核原則"),
        ("[16/17] 取得 Kernel hardening...", get_kernel_hardening, "KernelHardening"),
        ("[17/17] 讀取 hosts 檔案...", get_hosts_file, "Hosts檔案"),
    ]
    report_data: dict = {"meta": {"tool": "H2C_PcSecCheck_macos", "version": TOOL_VERSION, "schema_ver": SCHEMA_VERSION, "computer_name": computer_name, "local_ip": local_ip, "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}}
    for label, func, key in steps:
        _print(label)
        try:
            report_data[key] = func()
        except Exception as e:
            _print(f"  [錯誤] {key}: {e}")
            report_data[key] = []
    _print("\n[分析] 執行風險評估...")
    findings = analyze_findings(report_data)
    risk_score, risk_level = calculate_risk(findings)
    _print(f"[分析] 發現 {len(findings)} 個風險項目，等級: {risk_level.upper()} ({risk_score}/100)")
    _print("[打包] 產生 .h2cpc.zip...")
    zip_file = package_zip(base_name, report_data, findings, risk_score, risk_level, computer_name, local_ip)
    _print("")
    _print("=" * 56)
    _print(f"  完成！輸出：{zip_file}")
    _print(f"  風險等級：{risk_level.upper()} ({risk_score}/100)")
    _print("=" * 56)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[中斷] 使用者取消執行", flush=True)
    except Exception as e:
        print(f"\n[錯誤] {e}", flush=True)
        import traceback
        traceback.print_exc()
    input("\n按 Enter 鍵關閉視窗...")
