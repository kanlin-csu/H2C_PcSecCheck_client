#!/usr/bin/env python3
"""
H2C_PcSecCheck_linux v2.0
Linux PC 資安健診工具 — 自動收集系統資訊並產生標準化報告包（.h2cpc.zip）

Copyright 2026 H2C工作室 甘霖老師

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""
from __future__ import annotations

import grp
import html
import io
import os
import re
import socket
import subprocess
import json
import platform
import zipfile
import hashlib
import datetime
from openpyxl import Workbook

TOOL_VERSION = "2.0.0"
SCHEMA_VERSION = "2"

# ─────────────────────────────────────────────────────────────────────────────
# 風險規則常數（Linux 路徑版）
# ─────────────────────────────────────────────────────────────────────────────
_SUSPICIOUS_PATHS = re.compile(
    r"/(tmp|var/tmp|dev/shm|home/[^/]+/downloads|home/[^/]+/\.local/share/trash)/",
    re.IGNORECASE,
)
_SUSPICIOUS_CMDLINE = (
    ("base64 -d",       3),
    ("bash -i",         4),
    ("/dev/tcp/",       4),
    ("nc -e",           4),
    ("ncat -e",         4),
    ("mkfifo",          3),
    (r"curl.*\| *sh",   3),
    (r"wget.*\| *sh",   3),
    ("python.*-c.*exec",2),
    ("perl.*-e",        2),
)

_PRIVATE_NETS = (
    re.compile(r"^10\."),
    re.compile(r"^172\.(1[6-9]|2\d|3[01])\."),
    re.compile(r"^192\.168\."),
    re.compile(r"^127\."),
    re.compile(r"^::1$"),
    re.compile(r"^fe80:", re.I),
    re.compile(r"^0\.0\.0\.0$"),
)
_MALICIOUS_PORTS  = {4444, 5555, 6666, 1337, 31337, 9001}
# 23=Telnet(危險), 512-514=rsh/rexec/rlogin, 6379=Redis無auth, 27017=MongoDB無auth
_SENSITIVE_LISTEN = {23, 512, 513, 514, 445, 3306, 5432, 6379, 27017}

SEVERITY_LABEL = {4: "Critical", 3: "High", 2: "Medium", 1: "Low", 0: "Info"}
SEVERITY_COLOR = {
    4: "#dc3545", 3: "#fd7e14", 2: "#ffc107", 1: "#0dcaf0", 0: "#6c757d",
}
RISK_COLORS = {
    "critical": "#dc3545", "high": "#fd7e14",
    "medium":   "#ffc107", "low":  "#198754",
}


# ─────────────────────────────────────────────────────────────────────────────
# 工具函式
# ─────────────────────────────────────────────────────────────────────────────
def _run(cmd: list[str], timeout: int = 15, stdin_input: str | None = None) -> str:
    """執行系統指令，回傳 stdout 字串；失敗回傳空字串。"""
    try:
        r = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout,
            input=stdin_input,
        )
        return r.stdout or ""
    except FileNotFoundError:
        return ""
    except Exception:
        return ""


def _is_private_ip(ip: str) -> bool:
    ip = ip.split("%")[0]  # 去掉 IPv6 zone ID
    return any(p.match(ip) for p in _PRIVATE_NETS)


def get_local_ip_address() -> str:
    # 方法一：問核心路由表，不需要真的連上外網；無預設路由時可能失敗。
    out = _run(["ip", "-4", "route", "get", "1.1.1.1"], timeout=5)
    m = re.search(r"\bsrc\s+(\d+\.\d+\.\d+\.\d+)\b", out)
    if m and not m.group(1).startswith("127."):
        return m.group(1)

    # 方法二：hostname -I，適合沒有外網但有內網 IP 的主機。
    out = _run(["hostname", "-I"], timeout=5)
    for token in out.split():
        if re.match(r"^\d+\.\d+\.\d+\.\d+$", token) and not token.startswith("127."):
            return token

    # 方法三：掃描 ip addr 的 IPv4 位址。
    out = _run(["ip", "-4", "addr", "show", "scope", "global"], timeout=5)
    for line in out.splitlines():
        m = re.search(r"\binet\s+(\d+\.\d+\.\d+\.\d+)/", line)
        if m and not m.group(1).startswith("127."):
            return m.group(1)

    # 方法四：建立 UDP socket 取得本機來源 IP。
    try:
        with socket.create_connection(("8.8.8.8", 80), timeout=3) as s:
            ip = s.getsockname()[0]
            if ip and not ip.startswith("127."):
                return ip
    except Exception:
        pass

    # 方法五：hostname DNS 解析。
    try:
        for addr_info in socket.getaddrinfo(socket.gethostname(), None):
            ip = addr_info[4][0]
            if ":" not in ip and not ip.startswith("127."):
                return ip
    except Exception:
        pass
    return "未知"


# ─────────────────────────────────────────────────────────────────────────────
# 資料收集
# ─────────────────────────────────────────────────────────────────────────────
def get_system_info() -> list[dict]:
    data: list[dict] = [
        {"項目": "電腦名稱", "內容": platform.node()},
        {"項目": "區網 IP",  "內容": get_local_ip_address()},
    ]

    # /etc/os-release
    try:
        with open("/etc/os-release") as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    data.append({"項目": k.strip(), "內容": v.strip().strip('"')})
    except Exception:
        pass

    # uname
    uname = _run(["uname", "-a"]).strip()
    if uname:
        data.append({"項目": "uname -a", "內容": uname})

    # 核心版本
    data.append({"項目": "Kernel", "內容": platform.release()})
    data.append({"項目": "架構",   "內容": platform.machine()})

    # CPU 核心數
    nproc = _run(["nproc"]).strip()
    if nproc:
        data.append({"項目": "CPU 核心數", "內容": nproc})

    # 記憶體
    try:
        with open("/proc/meminfo") as f:
            for line in f:
                if line.startswith(("MemTotal:", "MemAvailable:", "SwapTotal:")):
                    k, v = line.split(":", 1)
                    data.append({"項目": k.strip(), "內容": v.strip()})
    except Exception:
        pass

    # 運行時間
    uptime = _run(["uptime", "-p"]).strip()
    if uptime:
        data.append({"項目": "運行時間", "內容": uptime})

    # 時區
    tz = _run(["timedatectl", "show", "--property=Timezone", "--value"]).strip()
    if tz:
        data.append({"項目": "時區", "內容": tz})

    return data


def get_av_status() -> list[dict]:
    """偵測已安裝的防毒 / 端點保護工具。"""
    tools: list[dict] = []

    checks = [
        (["clamscan", "--version"],   "ClamAV"),
        (["rkhunter", "--version"],   "rkhunter"),
        (["chkrootkit", "-V"],        "chkrootkit"),
        (["aide", "--version"],       "AIDE (檔案完整性)"),
        (["tripwire", "--version"],   "Tripwire"),
        (["lynis", "show", "version"],"Lynis"),
    ]
    for cmd, name in checks:
        out = _run(cmd, timeout=5)
        if out.strip():
            # 嘗試取版本號第一行
            ver = out.strip().splitlines()[0][:80]
            # 檢查 systemd 服務狀態
            svc_map = {"ClamAV": "clamav-daemon", "rkhunter": "rkhunter"}
            svc_status = ""
            if name in svc_map:
                svc_out = _run(["systemctl", "is-active", svc_map[name]], timeout=5).strip()
                svc_status = "啟用（running）" if svc_out == "active" else f"停用（{svc_out}）"
            tools.append({
                "工具":     name,
                "版本":     ver,
                "狀態":     "已安裝",
                "即時保護": svc_status or "N/A",
            })

    if not tools:
        tools.append({
            "工具":     "未偵測到防毒工具",
            "版本":     "",
            "狀態":     "未安裝",
            "即時保護": "無",
        })
    return tools


def get_installed_packages() -> list[dict]:
    """取得已安裝套件清單（支援 dpkg / rpm）。"""
    # Debian / Ubuntu
    out = _run(["dpkg", "-l"], timeout=30)
    if out:
        pkgs = []
        for line in out.splitlines():
            if line.startswith("ii "):
                parts = line.split(None, 4)
                if len(parts) >= 4:
                    pkgs.append({
                        "名稱": parts[1],
                        "版本": parts[2],
                        "架構": parts[3],
                        "描述": parts[4].strip() if len(parts) > 4 else "",
                    })
        if pkgs:
            return pkgs

    # RedHat / CentOS / Fedora
    out = _run(
        ["rpm", "-qa", "--queryformat", "%{NAME}\t%{VERSION}-%{RELEASE}\t%{ARCH}\t%{SUMMARY}\n"],
        timeout=30,
    )
    if out:
        pkgs = []
        for line in out.splitlines():
            parts = line.split("\t", 3)
            if len(parts) >= 2:
                pkgs.append({
                    "名稱": parts[0],
                    "版本": parts[1] if len(parts) > 1 else "",
                    "架構": parts[2] if len(parts) > 2 else "",
                    "描述": parts[3] if len(parts) > 3 else "",
                })
        if pkgs:
            return pkgs

    return [{"名稱": "無法取得套件清單（dpkg / rpm 均不可用）", "版本": "", "架構": "", "描述": ""}]


def get_local_user_accounts() -> list[dict]:
    result: list[dict] = []
    no_login_shells = {
        "/bin/false", "/usr/sbin/nologin", "/sbin/nologin",
        "/usr/bin/nologin", "/bin/nologin",
    }

    try:
        with open("/etc/passwd") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split(":", 6)
                if len(parts) < 7:
                    continue
                username, _, uid_s, gid_s, desc, _, shell = parts
                uid = int(uid_s) if uid_s.isdigit() else -1
                gid = int(gid_s) if gid_s.isdigit() else -1
                enabled = shell not in no_login_shells

                # 群組名稱
                try:
                    group_name = grp.getgrgid(gid).gr_name
                except Exception:
                    group_name = str(gid)

                # 上次登入（lastlog）
                last_logon = ""
                out = _run(["lastlog", "-u", username], timeout=5)
                if out:
                    lines = out.strip().splitlines()
                    if len(lines) >= 2:
                        ll = lines[1]
                        if "Never logged in" in ll or "**Never logged in**" in ll:
                            last_logon = "從未登入"
                        else:
                            # lastlog 格式：username  pts/0  ip  Mon Jan  1 00:00:00 +0000 2024
                            tokens = ll.split()
                            if len(tokens) >= 4:
                                last_logon = " ".join(tokens[-(min(5, len(tokens))):])

                # 密碼到期（chage）
                pw_expires = "無法取得"
                out2 = _run(["chage", "-l", username], timeout=5)
                if out2:
                    for cl in out2.splitlines():
                        if "Password expires" in cl:
                            val = cl.split(":", 1)[1].strip()
                            pw_expires = "永不到期" if val.lower() == "never" else val
                            break

                result.append({
                    "帳號名稱":     username,
                    "UID":          uid,
                    "是否啟用":     "啟用" if enabled else "停用",
                    "帳號類型":     "一般使用者" if uid >= 1000 else "系統帳號",
                    "描述":         desc,
                    "Shell":        shell,
                    "上次登入":     last_logon,
                    "密碼到期":     pw_expires,
                    "所屬群組":     group_name,
                })
    except Exception as e:
        result.append({
            "帳號名稱": f"無法取得: {e}", "UID": "", "是否啟用": "",
            "帳號類型": "", "描述": "", "Shell": "", "上次登入": "", "密碼到期": "", "所屬群組": "",
        })
    return result


def get_password_policy() -> list[dict]:
    data: list[dict] = []

    # /etc/login.defs
    try:
        with open("/etc/login.defs") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split(None, 1)
                if len(parts) == 2:
                    data.append({"設定": parts[0], "值": parts[1]})
    except Exception as e:
        data.append({"設定": "無法讀取 /etc/login.defs", "值": str(e)})

    # pam pwquality
    for pam_path in ["/etc/security/pwquality.conf"]:
        if not os.path.isfile(pam_path):
            continue
        try:
            with open(pam_path) as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "=" in line:
                        k, v = line.split("=", 1)
                        data.append({"設定": f"[pwquality] {k.strip()}", "值": v.strip()})
        except Exception:
            pass

    return data if data else [{"設定": "無法取得密碼原則", "值": ""}]


def get_update_status() -> list[dict]:
    """列出可更新套件數量與前幾筆套件，供健診摘要快速判讀。"""
    result: list[dict] = []

    out = _run(["apt", "list", "--upgradable"], timeout=30)
    if out:
        rows = [line for line in out.splitlines() if "/" in line and not line.lower().startswith("listing")]
        result.append({"項目": "套件管理器", "內容": "apt"})
        result.append({"項目": "可更新套件數", "內容": str(len(rows))})
        for line in rows[:30]:
            result.append({"項目": "可更新套件", "內容": line})
        return result

    out = _run(["dnf", "check-update"], timeout=30)
    if out:
        rows = [line for line in out.splitlines() if line and not line.startswith(("Last metadata", "Obsoleting"))]
        result.append({"項目": "套件管理器", "內容": "dnf"})
        result.append({"項目": "可更新套件數", "內容": str(len(rows))})
        for line in rows[:30]:
            result.append({"項目": "可更新套件", "內容": line})
        return result

    out = _run(["yum", "check-update"], timeout=30)
    if out:
        rows = [line for line in out.splitlines() if line and not line.startswith(("Loaded plugins", "Obsoleting"))]
        result.append({"項目": "套件管理器", "內容": "yum"})
        result.append({"項目": "可更新套件數", "內容": str(len(rows))})
        for line in rows[:30]:
            result.append({"項目": "可更新套件", "內容": line})
        return result

    return [{"項目": "更新狀態", "內容": "無法取得（apt/dnf/yum 不可用或逾時）"}]


def get_privileged_accounts() -> list[dict]:
    """列出 UID 0、sudo/wheel/admin 群組帳號，這是健診必看項。"""
    result: list[dict] = []
    passwd_users: dict[str, dict] = {}
    try:
        with open("/etc/passwd") as f:
            for line in f:
                parts = line.strip().split(":")
                if len(parts) >= 7:
                    passwd_users[parts[0]] = {
                        "UID": int(parts[2]) if parts[2].isdigit() else -1,
                        "Shell": parts[6],
                    }
    except Exception:
        pass

    for name, info in passwd_users.items():
        if info["UID"] == 0:
            result.append({"帳號": name, "權限來源": "UID 0", "UID": info["UID"], "Shell": info["Shell"]})

    for group_name in ("sudo", "wheel", "admin"):
        try:
            g = grp.getgrnam(group_name)
            for member in g.gr_mem:
                info = passwd_users.get(member, {})
                result.append({
                    "帳號": member,
                    "權限來源": group_name,
                    "UID": info.get("UID", ""),
                    "Shell": info.get("Shell", ""),
                })
        except Exception:
            continue

    seen = set()
    dedup = []
    for row in result:
        key = (row.get("帳號"), row.get("權限來源"))
        if key not in seen:
            seen.add(key)
            dedup.append(row)
    return dedup if dedup else [{"帳號": "未偵測到 sudo/wheel/admin 或額外 UID 0 帳號", "權限來源": "", "UID": "", "Shell": ""}]


def get_ssh_security() -> list[dict]:
    """讀取 SSH 伺服器有效設定；sshd -T 可解析 Include 與預設值。"""
    result: list[dict] = []
    keys = {
        "port": "連接埠",
        "permitrootlogin": "允許 root 登入",
        "passwordauthentication": "允許密碼登入",
        "pubkeyauthentication": "允許金鑰登入",
        "permitemptypasswords": "允許空密碼",
        "x11forwarding": "X11 Forwarding",
        "maxauthtries": "最大驗證嘗試次數",
        "clientaliveinterval": "閒置逾時秒數",
        "allowusers": "允許使用者",
        "allowgroups": "允許群組",
    }

    out = _run(["sshd", "-T"], timeout=10)
    parsed: dict[str, str] = {}
    if out:
        for line in out.splitlines():
            parts = line.split(None, 1)
            if len(parts) == 2:
                parsed[parts[0].lower()] = parts[1].strip()
    else:
        for conf in ("/etc/ssh/sshd_config",):
            if os.path.isfile(conf):
                try:
                    with open(conf) as f:
                        for line in f:
                            line = line.strip()
                            if not line or line.startswith("#"):
                                continue
                            parts = line.split(None, 1)
                            if len(parts) == 2:
                                parsed[parts[0].lower()] = parts[1].strip()
                except Exception:
                    pass

    if not parsed:
        return [{"設定": "OpenSSH Server", "值": "未安裝或無法讀取 sshd_config", "狀態": "Info"}]

    for key, label in keys.items():
        result.append({"設定": label, "值": parsed.get(key, "未設定/預設"), "狀態": "待檢核"})
    return result


def get_kernel_hardening() -> list[dict]:
    checks = {
        "kernel.randomize_va_space": "ASLR",
        "net.ipv4.ip_forward": "IPv4 轉送",
        "net.ipv4.conf.all.accept_redirects": "接受 ICMP Redirect",
        "net.ipv4.conf.all.send_redirects": "送出 ICMP Redirect",
        "net.ipv4.conf.all.rp_filter": "反向路徑過濾",
        "net.ipv4.tcp_syncookies": "SYN Cookies",
        "kernel.kptr_restrict": "限制 kernel pointer 洩漏",
        "kernel.dmesg_restrict": "限制 dmesg",
    }
    result = []
    for key, label in checks.items():
        val = _run(["sysctl", "-n", key], timeout=5).strip()
        result.append({"項目": label, "sysctl": key, "值": val if val else "無法取得"})
    return result


def get_mac_status() -> list[dict]:
    result: list[dict] = []
    aa = _run(["aa-status"], timeout=10)
    if aa:
        first = aa.strip().splitlines()[0] if aa.strip().splitlines() else "AppArmor 已安裝"
        result.append({"機制": "AppArmor", "狀態": first, "詳細": aa[:400]})
    else:
        result.append({"機制": "AppArmor", "狀態": "未偵測或未啟用", "詳細": ""})

    se = _run(["getenforce"], timeout=5).strip()
    if se:
        result.append({"機制": "SELinux", "狀態": se, "詳細": ""})
    else:
        result.append({"機制": "SELinux", "狀態": "未偵測", "詳細": ""})
    return result


def get_container_status() -> list[dict]:
    result: list[dict] = []
    for runtime in ("docker", "podman"):
        out = _run([runtime, "ps", "--format", "{{.Names}}\t{{.Image}}\t{{.Status}}\t{{.Ports}}"], timeout=10)
        if out:
            for line in out.splitlines():
                parts = line.split("\t")
                result.append({
                    "Runtime": runtime,
                    "名稱": parts[0] if len(parts) > 0 else "",
                    "映像": parts[1] if len(parts) > 1 else "",
                    "狀態": parts[2] if len(parts) > 2 else "",
                    "Port": parts[3] if len(parts) > 3 else "",
                })
    return result if result else [{"Runtime": "未偵測到執行中容器", "名稱": "", "映像": "", "狀態": "", "Port": ""}]


def get_file_permission_risks() -> list[dict]:
    result: list[dict] = []
    for label, cmd in [
        ("SUID 檔案", ["find", "/usr", "/bin", "/sbin", "/opt", "-xdev", "-perm", "-4000", "-type", "f", "-print"]),
        ("SGID 檔案", ["find", "/usr", "/bin", "/sbin", "/opt", "-xdev", "-perm", "-2000", "-type", "f", "-print"]),
        ("全域可寫目錄", ["find", "/etc", "/var", "/opt", "-xdev", "-type", "d", "-perm", "-0002", "-print"]),
    ]:
        out = _run(cmd, timeout=20)
        for path in out.splitlines()[:80]:
            result.append({"類型": label, "路徑": path, "說明": "請確認是否符合最小權限原則"})
    return result if result else [{"類型": "未偵測到明顯檔案權限風險", "路徑": "", "說明": ""}]


def get_network_settings() -> list[dict]:
    result: list[dict] = []

    # DNS servers
    dns_servers: list[str] = []
    try:
        with open("/etc/resolv.conf") as f:
            for line in f:
                if line.startswith("nameserver"):
                    parts = line.split()
                    if len(parts) >= 2:
                        dns_servers.append(parts[1])
    except Exception:
        pass

    # ip -j addr show（JSON 格式，需較新版 iproute2）
    out = _run(["ip", "-j", "addr", "show"], timeout=10)
    if out.strip().startswith("["):
        try:
            interfaces = json.loads(out)
            for iface in interfaces:
                name = iface.get("ifname", "")
                if name == "lo":
                    continue
                ipv4 = [a["local"] for a in iface.get("addr_info", []) if a.get("family") == "inet"]
                pfx  = [f"/{a['prefixlen']}" for a in iface.get("addr_info", []) if a.get("family") == "inet"]
                result.append({
                    "介面名稱":   name,
                    "IP位址":     ", ".join(ipv4),
                    "子網前綴":   ", ".join(pfx),
                    "DNS server": ", ".join(dns_servers),
                })
            return result
        except Exception:
            pass

    # fallback：ip addr show（文字格式）
    out = _run(["ip", "addr", "show"], timeout=10)
    current: dict | None = None
    for line in out.splitlines():
        m = re.match(r"^\d+:\s+(\S+):", line)
        if m:
            name = m.group(1)
            if name != "lo":
                current = {"介面名稱": name, "IP位址": "", "子網前綴": "", "DNS server": ", ".join(dns_servers)}
                result.append(current)
        elif current:
            m2 = re.match(r"\s+inet\s+(\d[\d.]+)/(\d+)", line)
            if m2:
                current["IP位址"]   = m2.group(1)
                current["子網前綴"] = f"/{m2.group(2)}"

    return result if result else [{"介面名稱": "無法取得", "IP位址": "", "子網前綴": "", "DNS server": ""}]


def get_startup_items() -> list[dict]:
    data: list[dict] = []

    # systemd enabled 服務
    out = _run(["systemctl", "list-unit-files", "--state=enabled", "--no-legend", "--no-pager"], timeout=10)
    for line in out.splitlines():
        parts = line.split(None, 2)
        if not parts:
            continue
        unit = parts[0]
        # 取得單元檔路徑
        path_out = _run(["systemctl", "show", unit, "--property=FragmentPath"], timeout=5)
        path = ""
        for pl in path_out.splitlines():
            if pl.startswith("FragmentPath="):
                path = pl.split("=", 1)[1]
        data.append({"來源": "systemd", "名稱": unit, "狀態": "啟用", "命令/路徑": path})

    # crontab（root）
    out = _run(["crontab", "-l"], timeout=5)
    for line in out.splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            data.append({"來源": "crontab (root)", "名稱": line[:60], "狀態": "啟用", "命令/路徑": line})

    # /etc/cron.*
    for cron_dir in ["/etc/cron.d", "/etc/cron.daily", "/etc/cron.hourly", "/etc/cron.weekly", "/etc/cron.monthly"]:
        if os.path.isdir(cron_dir):
            for fname in sorted(os.listdir(cron_dir)):
                fpath = os.path.join(cron_dir, fname)
                if os.path.isfile(fpath):
                    data.append({"來源": cron_dir, "名稱": fname, "狀態": "啟用", "命令/路徑": fpath})

    # /etc/rc.local
    if os.path.isfile("/etc/rc.local"):
        try:
            with open("/etc/rc.local") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and line != "exit 0":
                        data.append({"來源": "/etc/rc.local", "名稱": line[:60], "狀態": "啟用", "命令/路徑": line})
        except Exception:
            pass

    # /etc/init.d/ 可執行腳本
    if os.path.isdir("/etc/init.d"):
        for fname in sorted(os.listdir("/etc/init.d")):
            fpath = os.path.join("/etc/init.d", fname)
            if os.path.isfile(fpath) and os.access(fpath, os.X_OK):
                data.append({"來源": "/etc/init.d", "名稱": fname, "狀態": "啟用", "命令/路徑": fpath})

    return data


def get_processes() -> list[dict]:
    result: list[dict] = []
    page_size = getattr(os, "sysconf", lambda _: 4096)("SC_PAGE_SIZE") if hasattr(os, "sysconf") else 4096

    try:
        for pid_str in os.listdir("/proc"):
            if not pid_str.isdigit():
                continue
            pid = int(pid_str)
            base = f"/proc/{pid}"
            try:
                name = open(f"{base}/comm").read().strip()
            except Exception:
                continue
            try:
                exe = os.readlink(f"{base}/exe")
            except Exception:
                exe = ""
            try:
                cmdline = open(f"{base}/cmdline", "rb").read().replace(b"\x00", b" ").decode("utf-8", errors="replace").strip()
            except Exception:
                cmdline = ""
            ppid = 0
            try:
                for sl in open(f"{base}/status"):
                    if sl.startswith("PPid:"):
                        ppid = int(sl.split()[1])
                        break
            except Exception:
                pass
            mem_kb = 0
            try:
                parts = open(f"{base}/statm").read().split()
                mem_kb = int(parts[1]) * page_size // 1024 if len(parts) > 1 else 0
            except Exception:
                pass
            result.append({
                "程序名稱":   name,
                "PID":        pid,
                "父PID":      ppid,
                "執行路徑":   exe,
                "命令列":     cmdline[:500],
                "記憶體(KB)": mem_kb,
            })
    except Exception as e:
        result.append({
            "程序名稱": f"無法讀取 /proc: {e}",
            "PID": 0, "父PID": 0, "執行路徑": "", "命令列": "", "記憶體(KB)": 0,
        })
    return result


def _get_package_owner(path: str, cache: dict[str, str]) -> str:
    """回傳檔案所屬套件，供管理平台白名單 / 基準比對使用。"""
    if not path or path in cache:
        return cache.get(path, "")

    owner = ""
    out = _run(["dpkg", "-S", path], timeout=3)
    if out and ":" in out:
        owner = out.split(":", 1)[0].strip()
    else:
        out = _run(["rpm", "-qf", path], timeout=3)
        if out and "is not owned" not in out and "not owned" not in out:
            owner = out.strip().splitlines()[0]
    cache[path] = owner
    return owner


def _empty_threat_fields() -> dict:
    """平台端後續填入 VirusTotal / MalwareBazaar / internal_ioc 結果。"""
    return {
        "vt_status": "not_queried",
        "threat_source": "",
        "first_seen": "",
        "last_seen": "",
    }


def get_process_hashes() -> list[dict]:
    result: list[dict] = []
    package_cache: dict[str, str] = {}
    try:
        for pid_str in os.listdir("/proc"):
            if not pid_str.isdigit():
                continue
            pid = int(pid_str)
            try:
                exe = os.readlink(f"/proc/{pid_str}/exe")
            except Exception:
                continue
            if not exe:
                continue

            is_deleted = "(deleted)" in exe
            clean_path = exe.replace(" (deleted)", "")
            proc_name = ""
            try:
                proc_name = open(f"/proc/{pid_str}/comm").read().strip()
            except Exception:
                pass

            row = {
                "程序名稱": proc_name,
                "PID": pid,
                "執行路徑": exe,
                "SHA256": "",
                "hash_status": "not_calculated",
                "套件來源": "",
                "是否已刪除": "是" if is_deleted else "否",
                **_empty_threat_fields(),
            }

            try:
                sha = hashlib.sha256()
                with open(clean_path, "rb") as f:
                    while True:
                        chunk = f.read(65536)
                        if not chunk:
                            break
                        sha.update(chunk)
                row["SHA256"] = sha.hexdigest().upper()
                row["hash_status"] = "ok"
                row["套件來源"] = _get_package_owner(clean_path, package_cache)
            except PermissionError:
                row["hash_status"] = "permission_denied"
            except Exception:
                row["hash_status"] = "failed"
            result.append(row)
    except Exception:
        pass
    return result


def get_netstat() -> list[dict]:
    result: list[dict] = []

    def _parse_ss(args: list[str], is_established: bool = False) -> list[dict]:
        rows = []
        out = _run(["ss"] + args, timeout=15)
        for line in out.splitlines()[1:]:
            parts = line.split()
            if len(parts) < 5:
                continue
            proto = parts[0].upper()
            state = parts[1].upper()
            local = parts[4]
            remote = parts[5] if len(parts) > 5 else "*"
            proc_name = ""
            if len(parts) >= 7:
                m = re.search(r'users:\(\("([^"]+)"', " ".join(parts))
                if m:
                    proc_name = m.group(1)
            pid_match = re.search(r'pid=(\d+)', " ".join(parts))
            pid_str = pid_match.group(1) if pid_match else ""
            if state in ("ESTAB", "ESTABLISHED"):
                state = "ESTABLISHED"
            rows.append({
                "協定":    proto,
                "本地位址": local,
                "遠端位址": remote,
                "狀態":    state,
                "PID":     pid_str,
                "程序名稱": proc_name,
            })
        return rows

    # 監聽中（TCP/UDP）
    result += _parse_ss(["-tulnp"])
    # ESTABLISHED 連線
    result += _parse_ss(["-tnp"])

    if not result:
        # fallback: netstat
        out = _run(["netstat", "-tulnp"], timeout=15)
        for line in out.splitlines():
            if not re.match(r"^(tcp|udp)", line, re.I):
                continue
            parts = line.split()
            if len(parts) < 6:
                continue
            proto  = parts[0].upper()
            local  = parts[3]
            remote = parts[4]
            state  = parts[5] if proto.startswith("TCP") and len(parts) > 5 else ""
            last   = parts[-1] if len(parts) >= 7 else ""
            pid_s, _, proc = last.partition("/")
            result.append({
                "協定":    proto,
                "本地位址": local,
                "遠端位址": remote,
                "狀態":    state,
                "PID":     pid_s,
                "程序名稱": proc,
            })

    return result


def get_firewall_status() -> list[dict]:
    result: list[dict] = []

    # ufw
    out = _run(["ufw", "status", "verbose"], timeout=10)
    if out.strip():
        status = "否"
        for line in out.splitlines():
            if line.lower().startswith("status:"):
                status = "是" if "active" in line.lower() else "否"
        result.append({
            "防火牆類型": "ufw",
            "設定檔":     "all",
            "啟用":       status,
            "詳細資訊":   out.strip()[:400],
        })
        return result

    # firewalld
    out = _run(["firewall-cmd", "--state"], timeout=10)
    if out.strip():
        enabled = "是" if "running" in out.lower() else "否"
        detail = _run(["firewall-cmd", "--list-all"], timeout=10)
        result.append({
            "防火牆類型": "firewalld",
            "設定檔":     "default",
            "啟用":       enabled,
            "詳細資訊":   detail.strip()[:400],
        })
        return result

    # iptables
    out = _run(["iptables", "-L", "-n", "--line-numbers"], timeout=10)
    if out.strip():
        has_rules = any(
            re.search(r"\b(ACCEPT|DROP|REJECT)\b", ln)
            for ln in out.splitlines()
            if not ln.startswith("Chain")
        )
        result.append({
            "防火牆類型": "iptables",
            "設定檔":     "filter",
            "啟用":       "是" if has_rules else "否（無規則）",
            "詳細資訊":   out.strip()[:400],
        })
        return result

    result.append({"防火牆類型": "未偵測到防火牆", "設定檔": "", "啟用": "否", "詳細資訊": ""})
    return result


def get_smb_status() -> list[dict]:
    result: list[dict] = []

    # Samba 版本
    ver_out = _run(["smbd", "--version"], timeout=5)
    if not ver_out.strip():
        ver_out = _run(["smbd", "-V"], timeout=5)
    if not ver_out.strip():
        result.append({"設定": "Samba", "值": "未安裝（smbd 不存在）"})
        return result

    result.append({"設定": "Samba 版本", "值": ver_out.strip().splitlines()[0]})

    # 服務狀態
    svc_out = _run(["systemctl", "is-active", "smbd"], timeout=5).strip()
    if not svc_out:
        svc_out = _run(["service", "smbd", "status"], timeout=5)
        svc_out = "active（running）" if "running" in svc_out.lower() else "inactive"
    result.append({"設定": "Samba 服務狀態", "值": svc_out})

    # 優先使用 testparm -sv 取得實際生效設定（含 include / 預設值）。
    effective = _run(["testparm", "-sv"], timeout=10, stdin_input="\n")
    if effective:
        for key in ("server min protocol", "client min protocol", "server max protocol"):
            m = re.search(rf"^\s*{re.escape(key)}\s*=\s*(.+)$", effective, re.M | re.I)
            if m:
                result.append({"設定": key, "值": m.group(1).strip()})

    # 讀取 smb.conf 作為備援與人工檢視
    for conf in ["/etc/samba/smb.conf", "/usr/local/samba/etc/smb.conf"]:
        if not os.path.isfile(conf):
            continue
        try:
            content = open(conf).read()
            m = re.search(r"^\s*(server\s+min\s+protocol|min\s+protocol)\s*=\s*(.+)$", content, re.M | re.I)
            if m:
                proto = m.group(2).strip().upper()
                result.append({"設定": "smb.conf min protocol", "值": proto})
            else:
                result.append({"設定": "smb.conf min protocol", "值": "未設定"})
        except Exception as e:
            result.append({"設定": f"無法讀取 {conf}", "值": str(e)})
        break

    return result


def get_shared_folders() -> list[dict]:
    result: list[dict] = []

    # Samba smb.conf
    for conf in ["/etc/samba/smb.conf"]:
        if not os.path.isfile(conf):
            continue
        try:
            current_share: str | None = None
            share_props: dict = {}
            with open(conf) as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith(("#", ";")):
                        continue
                    m = re.match(r"^\[(.+)\]$", line)
                    if m:
                        if current_share and current_share.lower() not in ("global", "printers", "print$", "homes"):
                            result.append({
                                "共用名稱": current_share,
                                "路徑":     share_props.get("path", ""),
                                "說明":     share_props.get("comment", ""),
                                "類型":     "Samba",
                            })
                        current_share = m.group(1)
                        share_props = {}
                    elif "=" in line and current_share:
                        k, v = line.split("=", 1)
                        share_props[k.strip().lower()] = v.strip()
            if current_share and current_share.lower() not in ("global", "printers", "print$", "homes"):
                result.append({
                    "共用名稱": current_share,
                    "路徑":     share_props.get("path", ""),
                    "說明":     share_props.get("comment", ""),
                    "類型":     "Samba",
                })
        except Exception as e:
            result.append({"共用名稱": f"無法讀取設定: {e}", "路徑": "", "說明": "", "類型": ""})

    # NFS exports
    if os.path.isfile("/etc/exports"):
        try:
            with open("/etc/exports") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = line.split(None, 1)
                    if parts:
                        result.append({
                            "共用名稱": os.path.basename(parts[0]) or parts[0],
                            "路徑":     parts[0],
                            "說明":     parts[1] if len(parts) > 1 else "",
                            "類型":     "NFS",
                        })
        except Exception:
            pass

    return result


def get_audit_policy() -> list[dict]:
    result: list[dict] = []

    # auditd 服務狀態
    svc = _run(["systemctl", "is-active", "auditd"], timeout=5).strip()
    result.append({"稽核類別": "auditd 服務狀態", "稽核設定": svc if svc else "未知（systemctl 不可用）"})

    # auditctl -l（列出規則）
    rules_out = _run(["auditctl", "-l"], timeout=10)
    if rules_out.strip():
        lines = rules_out.strip().splitlines()
        if lines and lines[0].strip() in ("-a never,task", "No rules"):
            result.append({"稽核類別": "稽核規則", "稽核設定": "無規則（No auditing）"})
        else:
            for line in lines:
                result.append({"稽核類別": "稽核規則", "稽核設定": line.strip()})
    else:
        result.append({"稽核類別": "auditctl", "稽核設定": "未安裝或無法執行（需 root）"})

    # 規則檔存在狀態
    for path in ["/etc/audit/audit.rules", "/etc/audit/rules.d"]:
        if os.path.isfile(path):
            result.append({"稽核類別": "稽核規則檔", "稽核設定": path})
        elif os.path.isdir(path):
            try:
                n = len(os.listdir(path))
                result.append({"稽核類別": "稽核規則目錄", "稽核設定": f"{path} （{n} 個檔案）"})
            except Exception:
                pass

    return result


def get_hosts_file() -> list[dict]:
    hosts_path = "/etc/hosts"
    result: list[dict] = []
    _STD_IPS  = {"127.0.0.1", "::1", "0.0.0.0", "127.0.1.1", "127.0.53.53",
                 "255.255.255.255", "ff02::1", "ff02::2"}
    _STD_HOST = {"localhost", "localhost.localdomain", "localhost6", "localhost6.localdomain6",
                 "ip6-localhost", "ip6-loopback", "ip6-allnodes", "ip6-allrouters",
                 "broadcasthost"}
    try:
        with open(hosts_path, encoding="utf-8", errors="replace") as f:
            for lineno, line in enumerate(f, 1):
                stripped = line.strip()
                if not stripped or stripped.startswith("#"):
                    continue
                parts = stripped.split()
                if len(parts) >= 2:
                    ip, hostname = parts[0], parts[1]
                    is_std = ip in _STD_IPS and hostname in _STD_HOST
                    result.append({
                        "行號":    lineno,
                        "IP":      ip,
                        "主機名稱": hostname,
                        "完整內容": stripped,
                        "是否標準": "是" if is_std else "否",
                    })
    except Exception as e:
        result.append({"行號": 0, "IP": "", "主機名稱": "", "完整內容": f"無法讀取: {e}", "是否標準": "否"})
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Findings 分析層
# ─────────────────────────────────────────────────────────────────────────────
def _finding(category: str, severity: int, title: str, detail: str, raw_data: dict | None = None) -> dict:
    return {
        "category": category,
        "severity": severity,
        "title":    title,
        "detail":   detail,
        "raw_data": raw_data or {},
        "status":   "open",
    }


def _analyze_processes(processes: list[dict], process_hashes: list[dict]) -> list[dict]:
    findings: list[dict] = []
    pid_map = {p["PID"]: (p.get("程序名稱") or "").lower() for p in processes if p.get("PID")}

    for proc in processes:
        name    = (proc.get("程序名稱") or "").lower()
        path    = (proc.get("執行路徑") or "").lower()
        cmdline = (proc.get("命令列")   or "").lower()
        pid     = proc.get("PID")
        ppid    = proc.get("父PID")
        ctx     = f"PID：{pid}　父PID：{ppid}"

        # 可疑路徑
        if path and _SUSPICIOUS_PATHS.search(path):
            findings.append(_finding(
                "suspicious_process", 3,
                f"程序執行於可疑路徑: {proc['程序名稱']}",
                f"{ctx}\n執行路徑: {proc['執行路徑']}",
                proc,
            ))

        # 可疑命令列
        for pattern, sev in _SUSPICIOUS_CMDLINE:
            if re.search(pattern, cmdline, re.I):
                findings.append(_finding(
                    "suspicious_process", sev,
                    f"程序命令列含可疑特徵: {proc['程序名稱']}",
                    f"{ctx}\n命令列符合規則 '{pattern}'\n{proc['命令列'][:300]}",
                    proc,
                ))
                break

    return findings


def _analyze_netstat(netstat: list[dict]) -> list[dict]:
    findings: list[dict] = []
    for conn in netstat:
        state   = conn.get("狀態",    "")
        proto   = conn.get("協定",    "")
        foreign = conn.get("遠端位址", "")
        local   = conn.get("本地位址", "")

        foreign_ip       = foreign.rsplit(":", 1)[0] if ":" in foreign else foreign
        foreign_port_str = foreign.rsplit(":", 1)[1] if ":" in foreign else "0"
        try:
            foreign_port = int(foreign_port_str)
        except Exception:
            foreign_port = 0

        local_ip_part  = local.rsplit(":", 1)[0] if ":" in local else ""
        local_port_str = local.rsplit(":", 1)[1] if ":" in local else "0"
        try:
            local_port = int(local_port_str)
        except Exception:
            local_port = 0

        if state == "ESTABLISHED" and foreign_ip and not _is_private_ip(foreign_ip):
            proc_name = conn.get("程序名稱", "")
            if foreign_port in _MALICIOUS_PORTS:
                findings.append(_finding(
                    "suspicious_connection", 3,
                    f"連線至已知惡意 Port {foreign_port}",
                    f"程序 {proc_name} 連線到 {foreign}",
                    conn,
                ))
            elif foreign_port not in (80, 443):
                findings.append(_finding(
                    "suspicious_connection", 2,
                    f"對外非標準 Port 連線：{proc_name} → {foreign}",
                    f"程序 {proc_name} 使用非標準 port {foreign_port} 連線至公網",
                    conn,
                ))

        _is_wildcard = local_ip_part in ("0.0.0.0", "::", "[::]", "*", "")
        if state in ("LISTEN", "LISTENING") and _is_wildcard and local_port in _SENSITIVE_LISTEN:
            port_names = {
                23: "Telnet（明文，危險）", 445: "SMB",
                512: "rexec", 513: "rlogin", 514: "rsh",
                3306: "MySQL", 5432: "PostgreSQL",
                6379: "Redis（可能無認證）", 27017: "MongoDB（可能無認證）",
            }
            findings.append(_finding(
                "open_port", 2 if local_port not in (23, 512, 513, 514) else 3,
                f"敏感服務對外監聽: {port_names.get(local_port, str(local_port))}",
                f"Port {local_port} 在 {local_ip_part or '*'} 對外監聽",
                conn,
            ))
    return findings


def _analyze_accounts(user_accounts: list[dict]) -> list[dict]:
    findings: list[dict] = []
    now = datetime.datetime.now()

    # 找所有 UID 0 帳號（root 之外）
    uid0_non_root = [a for a in user_accounts if a.get("UID") == 0 and a.get("帳號名稱") != "root"]
    for acc in uid0_non_root:
        findings.append(_finding(
            "account_anomaly", 4,
            f"非 root 帳號具有 UID 0：{acc['帳號名稱']}",
            f"帳號 {acc['帳號名稱']} 的 UID=0，擁有等同 root 的完整權限，極度危險。",
            acc,
        ))

    for acc in user_accounts:
        name    = acc.get("帳號名稱", "")
        enabled = acc.get("是否啟用") == "啟用"
        shell   = acc.get("Shell", "")
        uid     = acc.get("UID")
        account_type = acc.get("帳號類型", "")
        if not enabled:
            continue

        pw_exp = acc.get("密碼到期", "")
        last   = acc.get("上次登入", "")

        is_interactive = (
            name == "root"
            or account_type == "一般使用者"
            or (isinstance(uid, int) and uid >= 1000)
        )
        if is_interactive and shell not in ("/bin/false", "/usr/sbin/nologin", "/sbin/nologin", "/usr/bin/nologin"):
            if pw_exp in ("永不到期", "never", "Never", ""):
                findings.append(_finding(
                    "account_anomaly", 1,
                    f"可互動帳號密碼永不到期: {name}",
                    f"帳號 {name} 可互動登入且密碼設定為永不到期，請確認是否符合帳號管理政策。",
                    acc,
                ))

        if account_type == "一般使用者" and name not in ("nobody",):
            findings.append(_finding(
                "account_inventory", 0,
                f"非系統帳號：{name}",
                f"UID={uid}，Shell={shell}。此類帳號應列入人員帳號盤點。",
                acc,
            ))

        # 長期未登入（> 180 天）
        if last and last not in ("", "從未登入"):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%a %b %d %H:%M:%S %z %Y", "%a %b %d %H:%M:%S %Y"):
                try:
                    dt = datetime.datetime.strptime(last.strip(), fmt)
                    if dt.tzinfo:
                        dt = dt.replace(tzinfo=None)
                    days = (now - dt).days
                    if days > 180:
                        findings.append(_finding(
                            "account_anomaly", 1,
                            f"帳號長期未登入: {name}",
                            f"最後登入 {last}（{days} 天前）",
                            acc,
                        ))
                    break
                except Exception:
                    continue

    return findings


def _analyze_password_policy(password_policy: list[dict]) -> list[dict]:
    findings: list[dict] = []
    policy = {item["設定"]: item["值"] for item in password_policy}

    def _int(key: str) -> int | None:
        v = str(policy.get(key, "")).strip()
        if v.upper() in ("UNLIMITED", "NEVER", "-1", "0", ""):
            return 0 if v not in ("", "UNLIMITED") else None
        try:
            return int(v)
        except Exception:
            return None

    # PASS_MIN_LEN
    min_len = _int("PASS_MIN_LEN")
    if min_len is not None and min_len < 8:
        findings.append(_finding(
            "password_policy", 2,
            "密碼最短長度不足（/etc/login.defs）",
            f"PASS_MIN_LEN = {min_len}，建議至少 8 字元",
            {"PASS_MIN_LEN": min_len},
        ))

    # PASS_MAX_DAYS
    max_days_str = policy.get("PASS_MAX_DAYS", "")
    try:
        max_days = int(max_days_str)
        if max_days == 99999 or max_days <= 0:
            findings.append(_finding(
                "password_policy", 2,
                "密碼永不過期（/etc/login.defs）",
                f"PASS_MAX_DAYS = {max_days}，密碼最長使用期限實際上為永不過期",
                {"PASS_MAX_DAYS": max_days},
            ))
    except Exception:
        pass

    return findings


def _analyze_antivirus(av_status: list[dict]) -> list[dict]:
    findings: list[dict] = []
    if not av_status:
        return findings
    if av_status[0].get("狀態") == "未安裝":
        findings.append(_finding(
            "endpoint_protection", 3,
            "未偵測到任何防毒 / 端點保護工具",
            "未找到 ClamAV、rkhunter、chkrootkit 等工具，建議安裝端點保護軟體。",
            {},
        ))
        return findings
    for av in av_status:
        if av.get("即時保護") and "停用" in av.get("即時保護", ""):
            findings.append(_finding(
                "endpoint_protection", 2,
                f"{av['工具']} 已安裝但即時保護停用",
                f"{av['工具']} daemon 未在執行中，即時保護未啟用。",
                av,
            ))
    return findings


def _analyze_startup_items(startup_items: list[dict]) -> list[dict]:
    findings: list[dict] = []
    for item in startup_items:
        cmd    = (item.get("命令/路徑") or "").lower()
        name   = item.get("名稱", "")
        source = item.get("來源", "")
        if cmd and _SUSPICIOUS_PATHS.search(cmd):
            findings.append(_finding(
                "suspicious_startup", 3,
                f"啟動項目執行於可疑路徑: {name}",
                f"來源: {source}\n命令: {item['命令/路徑']}",
                item,
            ))
        for pattern, sev in _SUSPICIOUS_CMDLINE:
            if re.search(pattern, cmd, re.I):
                findings.append(_finding(
                    "suspicious_startup", sev,
                    f"啟動項目命令含可疑特徵: {name}",
                    f"來源: {source}\n命令: {item.get('命令/路徑','')[:200]}",
                    item,
                ))
                break
    return findings


def _analyze_firewall(firewall_data: list[dict]) -> list[dict]:
    findings: list[dict] = []
    for profile in firewall_data:
        if profile.get("防火牆類型") == "未偵測到防火牆":
            findings.append(_finding(
                "firewall", 3,
                "未偵測到任何防火牆",
                "未找到 ufw / firewalld / iptables，主機缺乏網路層防護。",
                profile,
            ))
        elif profile.get("啟用") in ("否", "否（無規則）"):
            fw_type = profile.get("防火牆類型", "")
            findings.append(_finding(
                "firewall", 3,
                f"防火牆未啟用：{fw_type}",
                f"{fw_type} 未啟用或無規則，主機缺乏網路層保護。",
                profile,
            ))
    return findings


def _analyze_smb(smb_data: list[dict]) -> list[dict]:
    findings: list[dict] = []
    cfg = {item.get("設定", ""): str(item.get("值", "")) for item in smb_data}
    svc = cfg.get("Samba 服務狀態", "")
    if not smb_data or cfg.get("Samba", "").startswith("未安裝"):
        return findings

    if "active" not in svc.lower() and "running" not in svc.lower():
        findings.append(_finding(
            "network_share", 0,
            "Samba 已安裝但服務未啟用",
            f"smbd 狀態：{svc}。目前未對外提供 SMB 分享，僅列資訊。",
            {"Samba 服務狀態": svc},
        ))
        return findings

    proto = (
        cfg.get("server min protocol")
        or cfg.get("smb.conf min protocol")
        or cfg.get("min protocol")
        or ""
    ).upper()
    if proto in ("NT1", "LANMAN1", "CORE", "COREPLUS"):
        findings.append(_finding(
            "smb_risk", 4,
            "Samba SMBv1 已啟用",
            "SMBv1 存在 EternalBlue 等嚴重漏洞，建議在 smb.conf 加入：\nmin protocol = SMB2",
            {"server min protocol": proto},
        ))
    elif not proto or proto in ("未設定", "DEFAULT"):
        findings.append(_finding(
            "smb_risk", 2,
            "Samba 未明確限制最低 SMB 協定",
            "smbd 正在執行，但未明確設定 server min protocol，建議設定為 SMB2 或以上。",
            cfg,
        ))
    return findings


def _analyze_updates(update_status: list[dict]) -> list[dict]:
    findings: list[dict] = []
    data = {item.get("項目", ""): item.get("內容", "") for item in update_status}
    try:
        count = int(data.get("可更新套件數", "0"))
        if count >= 50:
            findings.append(_finding(
                "patch_management", 3,
                f"可更新套件過多：{count} 個",
                "系統存在大量尚未更新套件，建議安排弱點修補與版本更新。",
                {"可更新套件數": count},
            ))
        elif count > 0:
            findings.append(_finding(
                "patch_management", 2,
                f"存在可更新套件：{count} 個",
                "系統有尚未套用的套件更新，請確認是否包含安全性修補。",
                {"可更新套件數": count},
            ))
    except Exception:
        pass
    return findings


def _analyze_privileged_accounts(priv_accounts: list[dict]) -> list[dict]:
    findings: list[dict] = []
    for acc in priv_accounts:
        name = acc.get("帳號", "")
        source = acc.get("權限來源", "")
        if not source:
            continue
        if source == "UID 0" and name != "root":
            findings.append(_finding(
                "privileged_account", 4,
                f"非 root 帳號具有 UID 0：{name}",
                f"{name} 擁有等同 root 的完整權限，應立即確認必要性。",
                acc,
            ))
        elif source in ("sudo", "wheel", "admin"):
            findings.append(_finding(
                "privileged_account", 1,
                f"管理權限帳號：{name}",
                f"{name} 屬於 {source} 群組，應列入特權帳號盤點。",
                acc,
            ))
    return findings


def _analyze_ssh_security(ssh_rows: list[dict]) -> list[dict]:
    findings: list[dict] = []
    cfg = {item.get("設定", ""): str(item.get("值", "")).lower() for item in ssh_rows}
    if cfg.get("openssh server", "").startswith("未安裝"):
        return findings

    root_login = cfg.get("允許 root 登入", "")
    if root_login in ("yes", "without-password", "prohibit-password"):
        findings.append(_finding(
            "ssh_security", 3 if root_login == "yes" else 2,
            "SSH 允許 root 登入",
            f"PermitRootLogin = {root_login}，建議停用 root 直接登入，改用一般帳號加 sudo。",
            {"PermitRootLogin": root_login},
        ))

    pass_auth = cfg.get("允許密碼登入", "")
    if pass_auth == "yes":
        findings.append(_finding(
            "ssh_security", 2,
            "SSH 允許密碼登入",
            "PasswordAuthentication = yes，建議改用金鑰登入並搭配 MFA / VPN / IP 限制。",
            {"PasswordAuthentication": pass_auth},
        ))

    empty_pw = cfg.get("允許空密碼", "")
    if empty_pw == "yes":
        findings.append(_finding(
            "ssh_security", 4,
            "SSH 允許空密碼登入",
            "PermitEmptyPasswords = yes，這是高風險設定，應立即關閉。",
            {"PermitEmptyPasswords": empty_pw},
        ))
    return findings


def _analyze_kernel_hardening(rows: list[dict]) -> list[dict]:
    findings: list[dict] = []
    vals = {item.get("sysctl", ""): item.get("值", "") for item in rows}
    expected = {
        "kernel.randomize_va_space": ("2", 2, "ASLR 未完整啟用"),
        "net.ipv4.ip_forward": ("0", 2, "IPv4 轉送已啟用"),
        "net.ipv4.conf.all.accept_redirects": ("0", 2, "允許 ICMP Redirect"),
        "net.ipv4.conf.all.send_redirects": ("0", 2, "允許送出 ICMP Redirect"),
        "net.ipv4.tcp_syncookies": ("1", 2, "SYN Cookies 未啟用"),
        "kernel.kptr_restrict": ("1", 1, "kernel pointer 保護不足"),
        "kernel.dmesg_restrict": ("1", 1, "dmesg 未限制讀取"),
    }
    for key, (want, sev, title) in expected.items():
        val = str(vals.get(key, ""))
        if val and val != want:
            findings.append(_finding(
                "kernel_hardening", sev,
                title,
                f"{key} = {val}，建議值為 {want}。",
                {"sysctl": key, "值": val, "建議值": want},
            ))
    return findings


def _analyze_mac_status(rows: list[dict]) -> list[dict]:
    findings: list[dict] = []
    for row in rows:
        mechanism = row.get("機制", "")
        status = row.get("狀態", "")
        status_l = status.lower()
        if mechanism == "AppArmor" and ("not" in status_l or "未偵測" in status):
            findings.append(_finding(
                "mandatory_access_control", 1,
                "AppArmor 未啟用或未偵測",
                "建議依發行版支援狀況啟用 AppArmor 或 SELinux，提高程序隔離能力。",
                row,
            ))
        if mechanism == "SELinux" and status in ("Disabled", "Permissive"):
            findings.append(_finding(
                "mandatory_access_control", 1,
                f"SELinux 狀態：{status}",
                "建議評估是否可設為 Enforcing，或確認已由 AppArmor 提供等效保護。",
                row,
            ))
    return findings


def _analyze_file_permission_risks(rows: list[dict]) -> list[dict]:
    findings: list[dict] = []
    risky = [r for r in rows if r.get("路徑")]
    suid = [r for r in risky if r.get("類型") == "SUID 檔案"]
    ww = [r for r in risky if r.get("類型") == "全域可寫目錄"]
    if len(suid) > 30:
        findings.append(_finding(
            "file_permission", 2,
            f"SUID 檔案數量偏多：{len(suid)}",
            "SUID 檔案應定期盤點，避免被利用進行權限提升。",
            {"SUID 檔案數": len(suid), "前幾筆": suid[:10]},
        ))
    if ww:
        findings.append(_finding(
            "file_permission", 2,
            f"發現全域可寫目錄：{len(ww)} 個",
            "全域可寫目錄需確認 sticky bit 與用途，避免被植入或覆寫檔案。",
            {"全域可寫目錄": ww[:20]},
        ))
    return findings


def _analyze_shared_folders(shared_data: list[dict]) -> list[dict]:
    findings: list[dict] = []
    for share in shared_data:
        findings.append(_finding(
            "shared_folder", 2,
            f"偵測到共用資料夾：{share.get('共用名稱', '')}（{share.get('類型','')}）",
            f"路徑：{share.get('路徑','')}　請確認存取權限是否符合最小權限原則。",
            share,
        ))
    return findings


def _analyze_audit_policy(audit_data: list[dict]) -> list[dict]:
    findings: list[dict] = []
    audit_map = {item.get("稽核類別", ""): item.get("稽核設定", "") for item in audit_data}

    svc_status = audit_map.get("auditd 服務狀態", "")
    auditctl_missing = audit_map.get("auditctl") == "未安裝或無法執行（需 root）"
    if auditctl_missing:
        findings.append(_finding(
            "audit_policy", 2,
            "auditd 未安裝或無法執行",
            "未偵測到 auditd / auditctl，或未以 root 執行導致無法讀取稽核規則。",
            {},
        ))
        return findings

    if svc_status and svc_status not in ("active", "running"):
        findings.append(_finding(
            "audit_policy", 2,
            "auditd 服務未運行",
            f"auditd 狀態：{svc_status}。無法記錄系統稽核事件，出事時難以追查。",
            {"auditd": svc_status},
        ))

    rule_entries = [v for k, v in audit_map.items() if k == "稽核規則"]
    if any("無規則" in v or "No auditing" in v for v in rule_entries):
        findings.append(_finding(
            "audit_policy", 2,
            "auditd 無稽核規則",
            "auditd 已安裝但無設定任何稽核規則，等同未啟用稽核。\n"
            "建議安裝 auditd 規則集，例如：apt install auditd audispd-plugins",
            {},
        ))

    return findings


def _analyze_hosts_file(hosts_data: list[dict]) -> list[dict]:
    findings: list[dict] = []
    non_std = []
    local_hostname = platform.node()
    for h in hosts_data:
        if h.get("是否標準") != "否" or not h.get("IP"):
            continue
        # Debian / Ubuntu / Kali 常見：127.0.1.1 hostname，屬正常主機名解析。
        if h.get("IP") == "127.0.1.1" and h.get("主機名稱") == local_hostname:
            continue
        non_std.append(h)
    if non_std:
        lines = [f"  {h['IP']}\t{h['主機名稱']}" for h in non_std[:20]]
        findings.append(_finding(
            "hosts_tampering", 3,
            f"hosts 檔案含 {len(non_std)} 筆非標準記錄",
            "可能為 DNS 劫持或惡意重導向，請確認以下記錄是否合法：\n" + "\n".join(lines),
            {"非標準記錄": non_std},
        ))
    return findings


def analyze_findings(report_data: dict) -> list[dict]:
    findings: list[dict] = []
    findings += _analyze_updates(report_data.get("更新狀態", []))
    findings += _analyze_processes(
        report_data.get("執行中程序", []),
        report_data.get("程序雜湊",   []),
    )
    findings += _analyze_netstat(report_data.get("網路連線", []))
    findings += _analyze_accounts(report_data.get("使用者帳號", []))
    findings += _analyze_privileged_accounts(report_data.get("特權帳號", []))
    findings += _analyze_password_policy(report_data.get("密碼原則", []))
    findings += _analyze_ssh_security(report_data.get("SSH安全設定", []))
    findings += _analyze_antivirus(report_data.get("防毒工具", []))
    findings += _analyze_startup_items(report_data.get("啟動項目", []))
    findings += _analyze_firewall(report_data.get("防火牆狀態", []))
    findings += _analyze_smb(report_data.get("SMB狀態", []))
    findings += _analyze_shared_folders(report_data.get("網路分享", []))
    findings += _analyze_audit_policy(report_data.get("稽核原則", []))
    findings += _analyze_kernel_hardening(report_data.get("KernelHardening", []))
    findings += _analyze_mac_status(report_data.get("MAC保護", []))
    findings += _analyze_file_permission_risks(report_data.get("檔案權限風險", []))
    findings += _analyze_hosts_file(report_data.get("Hosts檔案", []))
    return findings


def calculate_risk(findings: list[dict]) -> tuple[int, str]:
    SEV_SCORE = {4: 40, 3: 20, 2: 10, 1: 3, 0: 0}
    score = min(100, sum(SEV_SCORE.get(f.get("severity", 0), 0) for f in findings))
    if   score >= 80: level = "critical"
    elif score >= 50: level = "high"
    elif score >= 20: level = "medium"
    else:             level = "low"
    return score, level


def build_compliance_summary(report_data: dict, findings: list[dict]) -> list[dict]:
    """行政院資安健診常看重點，置頂顯示並用顏色快速判讀。"""
    policy = {item.get("設定", ""): item.get("值", "") for item in report_data.get("密碼原則", [])}
    users = report_data.get("使用者帳號", [])
    priv = report_data.get("特權帳號", [])
    av = report_data.get("防毒工具", [])
    fw = report_data.get("防火牆狀態", [])
    ssh = {item.get("設定", ""): str(item.get("值", "")) for item in report_data.get("SSH安全設定", [])}
    audit = {item.get("稽核類別", ""): item.get("稽核設定", "") for item in report_data.get("稽核原則", [])}
    updates = {item.get("項目", ""): item.get("內容", "") for item in report_data.get("更新狀態", [])}
    shares = report_data.get("網路分享", []) or report_data.get("共用資料夾", [])
    listening = [
        n for n in report_data.get("網路連線", [])
        if str(n.get("狀態", "")).upper() in ("LISTEN", "LISTENING")
    ]

    def row(item: str, value: str, status: str, suggestion: str) -> dict:
        return {"檢核項目": item, "目前狀態": value, "判定": status, "建議": suggestion}

    summary: list[dict] = []

    min_len = policy.get("PASS_MIN_LEN", "未設定")
    try:
        min_len_i = int(min_len)
        summary.append(row(
            "密碼最短長度",
            str(min_len_i),
            "PASS" if min_len_i >= 8 else "FAIL",
            "建議至少 8 字元，若適用更高要求可提升至 12 字元。",
        ))
    except Exception:
        summary.append(row("密碼最短長度", str(min_len), "WARN", "未能確認 PASS_MIN_LEN，請檢查 PAM / login.defs。"))

    max_days = policy.get("PASS_MAX_DAYS", "未設定")
    try:
        max_days_i = int(max_days)
        summary.append(row(
            "密碼最長期限",
            str(max_days_i),
            "PASS" if 0 < max_days_i <= 180 else "WARN",
            "建議依機關政策設定密碼期限，避免 99999 或永不過期。",
        ))
    except Exception:
        summary.append(row("密碼最長期限", str(max_days), "WARN", "未能確認 PASS_MAX_DAYS。"))

    interactive = [
        u for u in users
        if (u.get("帳號類型") == "一般使用者" or u.get("帳號名稱") == "root")
        and u.get("是否啟用") == "啟用"
    ]
    summary.append(row(
        "非系統帳號 / 可互動帳號",
        f"{len(interactive)} 個",
        "INFO" if interactive else "PASS",
        "請確認是否皆有對應人員、用途與離職停用流程。",
    ))

    priv_count = len([p for p in priv if p.get("權限來源")])
    summary.append(row(
        "特權帳號",
        f"{priv_count} 個",
        "WARN" if priv_count > 1 else "PASS",
        "sudo / wheel / UID 0 帳號需定期盤點並留存授權紀錄。",
    ))

    av_missing = av and av[0].get("狀態") == "未安裝"
    summary.append(row(
        "防毒 / 端點防護",
        av[0].get("工具", "無資料") if av else "無資料",
        "WARN" if av_missing else "PASS",
        "Linux Server 可依政策採 EDR、ClamAV、完整性檢查或集中監控。",
    ))

    fw_bad = any(f.get("啟用") in ("否", "否（無規則）") or f.get("防火牆類型") == "未偵測到防火牆" for f in fw)
    summary.append(row(
        "主機防火牆",
        "; ".join(f"{f.get('防火牆類型')}={f.get('啟用')}" for f in fw) if fw else "無資料",
        "FAIL" if fw_bad else "PASS",
        "建議啟用 ufw / firewalld / nftables / iptables 並限制入站服務。",
    ))

    root_login = ssh.get("允許 root 登入", "未偵測")
    pass_auth = ssh.get("允許密碼登入", "未偵測")
    ssh_bad = root_login.lower() == "yes" or pass_auth.lower() == "yes"
    summary.append(row(
        "SSH 登入安全",
        f"RootLogin={root_login}, PasswordAuth={pass_auth}",
        "WARN" if ssh_bad else "PASS",
        "建議停用 root 直接登入，改用金鑰登入與 sudo 管理。",
    ))

    audit_state = audit.get("auditd 服務狀態", "未知")
    summary.append(row(
        "稽核記錄 auditd",
        audit_state,
        "PASS" if audit_state == "active" else "WARN",
        "建議啟用 auditd 並設定登入、權限提升、重要檔案變更等規則。",
    ))

    update_count = updates.get("可更新套件數", "未知")
    try:
        upd_i = int(update_count)
        upd_status = "PASS" if upd_i == 0 else "WARN"
    except Exception:
        upd_status = "INFO"
    summary.append(row(
        "修補 / 更新狀態",
        str(update_count),
        upd_status,
        "請確認可更新套件是否包含安全性修補，並保留更新紀錄。",
    ))

    summary.append(row(
        "對外監聽服務",
        f"{len(listening)} 個",
        "INFO" if listening else "PASS",
        "請確認每個 LISTEN 服務的業務必要性、來源限制與防火牆規則。",
    ))

    summary.append(row(
        "網路分享 / 對外目錄",
        f"{len(shares)} 筆",
        "WARN" if shares else "PASS",
        "若存在 Samba / NFS / Web / FTP 對外目錄，請確認最小權限與存取控制。",
    ))

    return summary


# ─────────────────────────────────────────────────────────────────────────────
# HTML 產生（與 Windows 版相同，自含式）
# ─────────────────────────────────────────────────────────────────────────────
_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Noto Sans TC', Arial, sans-serif; font-size: 13px;
       background: #f0f2f5; color: #212529; }
.topbar { background: #212529; color: #fff; padding: 10px 20px;
          position: sticky; top: 0; z-index: 100;
          display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
.topbar .brand { font-size: 15px; font-weight: 700; margin-right: 8px; }
.topbar a { color: #adb5bd; text-decoration: none; padding: 4px 8px;
            border-radius: 4px; font-size: 12px; }
.topbar a:hover { background: #343a40; color: #fff; }
.container { max-width: 1440px; margin: 0 auto; padding: 20px; }
h1 { font-size: 20px; margin: 16px 0 4px; }
.meta-line { color: #6c757d; margin-bottom: 20px; font-size: 12px; }
.card { background: #fff; border: 1px solid #dee2e6; border-radius: 6px;
        margin-bottom: 22px; }
.card-header { padding: 9px 16px; background: #343a40; color: #fff;
               border-radius: 6px 6px 0 0; font-weight: 600;
               display: flex; justify-content: space-between; align-items: center; }
.card-body { padding: 14px; }
.summary-grid { display: grid;
                grid-template-columns: repeat(auto-fill, minmax(120px, 1fr));
                gap: 10px; margin-bottom: 18px; }
.s-item { background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px;
          padding: 10px; text-align: center; }
.s-item .val { font-size: 26px; font-weight: 700; }
.s-item .lbl { font-size: 11px; color: #6c757d; margin-top: 3px; }
.badge { display: inline-block; padding: 3px 10px; border-radius: 12px;
         color: #fff; font-size: 11px; font-weight: 600; }
.search { margin-bottom: 8px; }
.search input { padding: 5px 10px; border: 1px solid #ced4da; border-radius: 4px;
                font-size: 12px; width: 100%; max-width: 300px; }
.tbl-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 12px; }
th { background: #495057; color: #fff; padding: 6px 8px; text-align: left;
     cursor: pointer; white-space: nowrap; user-select: none; }
th:hover { background: #6c757d; }
td { padding: 5px 8px; border-bottom: 1px solid #e9ecef;
     word-break: break-word; max-width: 420px; vertical-align: top; }
tr:hover td { background: #f8f9fa; }
.ns-ESTABLISHED { background: #d1e7dd; }
.ns-LISTEN, .ns-LISTENING { background: #cff4fc; }
.status { display:inline-block; min-width:56px; text-align:center; padding:3px 8px; border-radius:10px; color:#fff; font-weight:700; font-size:11px; }
.st-PASS { background:#198754; }
.st-WARN { background:#ffc107; color:#212529; }
.st-FAIL { background:#dc3545; }
.st-INFO { background:#0dcaf0; color:#212529; }
.focus-note { color:#6c757d; font-size:12px; margin-bottom:10px; }
"""

_JS = """
function sortT(id, col) {
  var t = document.getElementById(id); if (!t) return;
  var dir = t.dataset.dir === "asc" ? "desc" : "asc"; t.dataset.dir = dir;
  var rows = Array.from(t.querySelectorAll("tbody tr"));
  rows.sort(function(a, b) {
    var x = (a.cells[col] || {}).innerText || "";
    var y = (b.cells[col] || {}).innerText || "";
    var c = x.localeCompare(y, "zh-Hant", {numeric: true, sensitivity: "base"});
    return dir === "asc" ? c : -c;
  });
  var tb = t.querySelector("tbody");
  rows.forEach(function(r) { tb.appendChild(r); });
}
function filterT(inp, id) {
  var f = inp.value.toLowerCase();
  document.getElementById(id).querySelectorAll("tbody tr").forEach(function(r) {
    r.style.display = r.innerText.toLowerCase().includes(f) ? "" : "none";
  });
}
"""


def _section_table(data: list, tid: str) -> str:
    if not data:
        return '<p style="color:#6c757d;padding:8px">無資料</p>'
    if not (isinstance(data, list) and data and isinstance(data[0], dict)):
        return "".join(f"<p>{item}</p>" for item in data)
    headers = list(data[0].keys())
    h = [
        f'<div class="search"><input placeholder="搜尋..." oninput="filterT(this,\'{tid}\')"></div>',
        f'<div class="tbl-wrap"><table id="{tid}"><thead><tr>',
    ]
    for i, hdr in enumerate(headers):
        h.append(f'<th onclick="sortT(\'{tid}\',{i})">{hdr} ↕</th>')
    h.append("</tr></thead><tbody>")
    for item in data:
        h.append("<tr>")
        for hdr in headers:
            val = item.get(hdr, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            h.append(f"<td>{html.escape(str(val))}</td>")
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "\n".join(h)


def _netstat_table(data: list) -> str:
    if not data:
        return '<p style="color:#6c757d;padding:8px">無資料</p>'
    headers = ["協定", "本地位址", "遠端位址", "狀態", "PID", "程序名稱"]
    h = [
        '<div class="search"><input placeholder="搜尋..." oninput="filterT(this,\'ns_t\')"></div>',
        '<div class="tbl-wrap"><table id="ns_t"><thead><tr>',
    ]
    for i, hdr in enumerate(headers):
        h.append(f'<th onclick="sortT(\'ns_t\',{i})">{hdr} ↕</th>')
    h.append("</tr></thead><tbody>")
    for item in data:
        state = item.get("狀態", "")
        cls   = f"ns-{state}" if state in ("ESTABLISHED", "LISTEN", "LISTENING") else ""
        h.append(f'<tr class="{cls}">')
        for hdr in headers:
            h.append(f"<td>{html.escape(str(item.get(hdr, '')))}</td>")
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "\n".join(h)


def _focus_table(summary: list[dict]) -> str:
    headers = ["檢核項目", "目前狀態", "判定", "建議"]
    h = ['<p class="focus-note">以下為行政院資安健診常見關注項目，方便先看重點；詳細原始資料仍保留在下方各章節。</p>']
    h.append('<div class="tbl-wrap"><table id="focus_t"><thead><tr>')
    for i, hdr in enumerate(headers):
        h.append(f'<th onclick="sortT(\'focus_t\',{i})">{hdr} ↕</th>')
    h.append("</tr></thead><tbody>")
    for item in summary:
        status = str(item.get("判定", "INFO")).upper()
        h.append("<tr>")
        h.append(f"<td>{html.escape(str(item.get('檢核項目', '')))}</td>")
        h.append(f"<td>{html.escape(str(item.get('目前狀態', '')))}</td>")
        h.append(f'<td><span class="status st-{status}">{status}</span></td>')
        h.append(f"<td>{html.escape(str(item.get('建議', '')))}</td>")
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "\n".join(h)


def generate_html(computer_name: str, local_ip: str, report_data: dict,
                  findings: list, risk_score: int, risk_level: str) -> str:
    sc = {s: sum(1 for f in findings if f.get("severity") == s) for s in (4, 3, 2, 1, 0)}
    risk_color = RISK_COLORS.get(risk_level, "#6c757d")
    now_str    = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    focus_summary = build_compliance_summary(report_data, findings)

    def findings_rows() -> str:
        rows = []
        for f in sorted(findings, key=lambda x: -x.get("severity", 0)):
            sev    = f.get("severity", 0)
            color  = SEVERITY_COLOR.get(sev, "#6c757d")
            badge  = f'<span class="badge" style="background:{color}">{SEVERITY_LABEL.get(sev,"?")}</span>'
            detail = html.escape(f.get("detail") or "").replace("\n", "<br>")
            rows.append(
                f"<tr><td>{badge}</td>"
                f"<td>{html.escape(f.get('category',''))}</td>"
                f"<td>{html.escape(f.get('title',''))}</td>"
                f"<td>{detail}</td></tr>"
            )
        if not rows:
            return '<tr><td colspan="4" style="color:#6c757d;text-align:center">無風險項目</td></tr>'
        return "\n".join(rows)

    sections = [
        ("系統資訊",             "sys_t",       report_data.get("系統資訊",        [])),
        ("修補 / 更新狀態",       "upd_t",       report_data.get("更新狀態",        [])),
        ("帳號與權限",           "usr_t",       report_data.get("使用者帳號",      [])),
        ("特權帳號",             "priv_t",      report_data.get("特權帳號",        [])),
        ("密碼 / PAM 原則",       "pol_t",       report_data.get("密碼原則",        [])),
        ("SSH 安全設定",         "ssh_t",       report_data.get("SSH安全設定",     [])),
        ("防毒 / 端點防護",       "av_t",        report_data.get("防毒工具",        [])),
        ("防火牆狀態",           "fw_t",        report_data.get("防火牆狀態",      [])),
        ("網路設定",             "net_t",       report_data.get("網路設定",        [])),
        ("對外監聽 / 網路連線",   "conn_t",      report_data.get("網路連線",        [])),
        ("啟動服務與排程",       "startup_t",   report_data.get("啟動項目",        [])),
        ("稽核與日誌",           "audit_t",     report_data.get("稽核原則",        [])),
        ("Kernel Hardening",     "kernel_t",    report_data.get("KernelHardening", [])),
        ("SELinux / AppArmor",   "mac_t",       report_data.get("MAC保護",         [])),
        ("檔案權限風險",         "perm_t",      report_data.get("檔案權限風險",    [])),
        ("網路分享 / 對外目錄",   "share_t",     report_data.get("網路分享",        [])),
        ("Samba 設定",           "samba_t",     report_data.get("SMB狀態",         [])),
        ("容器環境",             "container_t", report_data.get("容器環境",        [])),
        ("Hosts 檔案",           "hosts_t",     report_data.get("Hosts檔案",       [])),
        ("已安裝套件",           "pkg_t",       report_data.get("已安裝套件",      [])),
        ("執行中程序",           "proc_t",      report_data.get("執行中程序",      [])),
        ("程序情資素材 / Hash 清單", "hash_t",   report_data.get("程序雜湊",        [])),
    ]
    nav = " ".join(
        f'<a href="#{aid}">{label}</a>'
        for label, aid in [("健診重點", "focus_sec"), ("風險摘要", "findings_sec")]
                          + [(t, f"sec_{i}") for t, i, _ in sections]
    )
    cards = ""
    for title, tid, data in sections:
        cards += f"""
<div class="card" id="sec_{tid}">
  <div class="card-header">{title}</div>
  <div class="card-body">{_section_table(data, tid)}</div>
</div>"""

    return f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>H2C Linux 健診 — {computer_name} ({local_ip})</title>
<style>{_CSS}</style>
</head>
<body>
<div class="topbar">
  <span class="brand">H2C Linux 健診</span>
  {nav}
</div>
<div class="container">
  <h1>Linux 資安健診報告</h1>
  <p class="meta-line">
    主機：<strong>{computer_name}</strong> &nbsp;|&nbsp;
    IP：<strong>{local_ip}</strong> &nbsp;|&nbsp;
    產生時間：{now_str}
  </p>
  <div class="card" id="focus_sec">
    <div class="card-header">行政院資安健診重點摘要</div>
    <div class="card-body">{_focus_table(focus_summary)}</div>
  </div>
  <div class="card" id="findings_sec">
    <div class="card-header">
      風險摘要
      <span class="badge" style="background:{risk_color};font-size:13px;padding:4px 14px">
        {risk_level.upper()} &nbsp; {risk_score}/100
      </span>
    </div>
    <div class="card-body">
      <div class="summary-grid">
        <div class="s-item"><div class="val" style="color:#dc3545">{sc[4]}</div><div class="lbl">Critical</div></div>
        <div class="s-item"><div class="val" style="color:#fd7e14">{sc[3]}</div><div class="lbl">High</div></div>
        <div class="s-item"><div class="val" style="color:#ffc107">{sc[2]}</div><div class="lbl">Medium</div></div>
        <div class="s-item"><div class="val" style="color:#0dcaf0">{sc[1]}</div><div class="lbl">Low</div></div>
        <div class="s-item"><div class="val" style="color:#6c757d">{sc[0]}</div><div class="lbl">Info</div></div>
      </div>
      <div class="tbl-wrap">
        <table id="findings_t">
          <thead><tr>
            <th onclick="sortT('findings_t',0)">嚴重度 ↕</th>
            <th onclick="sortT('findings_t',1)">類別 ↕</th>
            <th onclick="sortT('findings_t',2)">標題 ↕</th>
            <th>說明</th>
          </tr></thead>
          <tbody>{findings_rows()}</tbody>
        </table>
      </div>
    </div>
  </div>
{cards}
</div>
<script>{_JS}</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# XLSX 產生
# ─────────────────────────────────────────────────────────────────────────────
def _safe_cell(val):
    s = str(val) if not isinstance(val, (int, float, type(None))) else val
    if isinstance(s, str) and s and s[0] in ("=", "+", "-", "@", "|", "%"):
        return "'" + s
    return val


def _write_sheet(ws, data: list, field_order: list | None = None) -> None:
    if not data:
        ws.append(["無資料"])
        return
    if field_order is None:
        field_order = list(data[0].keys()) if data else []
    ws.append(field_order)
    for item in data:
        row = []
        for field in field_order:
            val = item.get(field, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            row.append(_safe_cell(val))
        ws.append(row)


def generate_xlsx(report_data: dict, findings: list) -> "Workbook":
    wb = Workbook()
    ws = wb.active
    ws.title = "健診重點"
    focus = build_compliance_summary(report_data, findings)
    _write_sheet(ws, focus, ["檢核項目", "目前狀態", "判定", "建議"])

    ws = wb.create_sheet("風險摘要")
    ws.append(["類別", "嚴重度", "標題", "說明", "狀態"])
    for f in sorted(findings, key=lambda x: -x.get("severity", 0)):
        ws.append([
            _safe_cell(f.get("category", "")),
            SEVERITY_LABEL.get(f.get("severity", 0), ""),
            _safe_cell(f.get("title", "")),
            _safe_cell(f.get("detail", "")),
            f.get("status", "open"),
        ])

    sheet_cfg = [
        ("系統資訊",         report_data.get("系統資訊",   []),  ["項目", "內容"]),
        ("更新狀態",         report_data.get("更新狀態",   []),  ["項目", "內容"]),
        ("防毒工具",         report_data.get("防毒工具",   []),  ["工具", "版本", "狀態", "即時保護"]),
        ("已安裝套件",       report_data.get("已安裝套件", []),  ["名稱", "版本", "架構", "描述"]),
        ("使用者帳號",       report_data.get("使用者帳號", []),
         ["帳號名稱", "UID", "是否啟用", "帳號類型", "描述", "Shell", "上次登入", "密碼到期", "所屬群組"]),
        ("特權帳號",         report_data.get("特權帳號",   []),  ["帳號", "權限來源", "UID", "Shell"]),
        ("密碼原則",         report_data.get("密碼原則",   []),  ["設定", "值"]),
        ("SSH安全設定",      report_data.get("SSH安全設定", []), ["設定", "值", "狀態"]),
        ("網路設定",         report_data.get("網路設定",   []),  ["介面名稱", "IP位址", "子網前綴", "DNS server"]),
        ("啟動項目",         report_data.get("啟動項目",   []),  ["來源", "名稱", "狀態", "命令/路徑"]),
        ("執行中程序",       report_data.get("執行中程序", []),
         ["程序名稱", "PID", "父PID", "執行路徑", "命令列", "記憶體(KB)"]),
        ("程序情資素材",     report_data.get("程序雜湊",   []),
         ["程序名稱", "PID", "執行路徑", "SHA256", "hash_status", "套件來源",
          "是否已刪除", "vt_status", "threat_source", "first_seen", "last_seen"]),
        ("網路連線",         report_data.get("網路連線",   []),  ["協定", "本地位址", "遠端位址", "狀態", "PID", "程序名稱"]),
        ("防火牆狀態",       report_data.get("防火牆狀態", []),  ["防火牆類型", "設定檔", "啟用", "詳細資訊"]),
        ("SMB狀態",          report_data.get("SMB狀態",    []),  ["設定", "值"]),
        ("網路分享",         report_data.get("網路分享",   []),  ["共用名稱", "路徑", "說明", "類型"]),
        ("稽核原則",         report_data.get("稽核原則",   []),  ["稽核類別", "稽核設定"]),
        ("KernelHardening",  report_data.get("KernelHardening", []), ["項目", "sysctl", "值"]),
        ("MAC保護",          report_data.get("MAC保護", []), ["機制", "狀態", "詳細"]),
        ("容器環境",         report_data.get("容器環境", []), ["Runtime", "名稱", "映像", "狀態", "Port"]),
        ("檔案權限風險",     report_data.get("檔案權限風險", []), ["類型", "路徑", "說明"]),
        ("Hosts檔案",        report_data.get("Hosts檔案",  []),  ["行號", "IP", "主機名稱", "完整內容", "是否標準"]),
    ]
    for sheet_name, data, fields in sheet_cfg:
        _write_sheet(wb.create_sheet(sheet_name), data, fields)

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# 打包 .h2cpc.zip
# ─────────────────────────────────────────────────────────────────────────────
def package_zip(base_name: str, report_data: dict, findings: list,
                risk_score: int, risk_level: str,
                computer_name: str, local_ip: str) -> str:
    now = datetime.datetime.now()
    report_json_bytes   = json.dumps(report_data, ensure_ascii=False, indent=2).encode("utf-8")
    findings_with_id    = [dict(f, id=f"linux:{f['category']}:{i}") for i, f in enumerate(findings)]
    findings_json_bytes = json.dumps(findings_with_id, ensure_ascii=False, indent=2).encode("utf-8")
    html_bytes          = generate_html(computer_name, local_ip, report_data, findings, risk_score, risk_level).encode("utf-8")
    buf = io.BytesIO()
    generate_xlsx(report_data, findings).save(buf)
    xlsx_bytes = buf.getvalue()

    meta = {
        "tool":          "H2C_PcSecCheck_linux",
        "version":       TOOL_VERSION,
        "schema_ver":    SCHEMA_VERSION,
        "computer_name": computer_name,
        "local_ip":      local_ip,
        "generated_at":  now.strftime("%Y-%m-%d %H:%M:%S"),
        "risk_score":    risk_score,
        "risk_level":    risk_level,
        "finding_count": {
            "critical": sum(1 for f in findings if f.get("severity") == 4),
            "high":     sum(1 for f in findings if f.get("severity") == 3),
            "medium":   sum(1 for f in findings if f.get("severity") == 2),
            "low":      sum(1 for f in findings if f.get("severity") == 1),
            "info":     sum(1 for f in findings if f.get("severity") == 0),
        },
        "checksums": {
            f"{base_name}.report.json":   hashlib.sha256(report_json_bytes).hexdigest(),
            f"{base_name}.findings.json": hashlib.sha256(findings_json_bytes).hexdigest(),
        },
    }
    meta_json_bytes = json.dumps(meta, ensure_ascii=False, indent=2).encode("utf-8")
    zip_filename = f"{base_name}.h2cpc.zip"
    with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("meta.json",                    meta_json_bytes)
        zf.writestr(f"{base_name}.report.json",     report_json_bytes)
        zf.writestr(f"{base_name}.findings.json",   findings_json_bytes)
        zf.writestr(f"{base_name}.report.html",     html_bytes)
        zf.writestr(f"{base_name}.report.xlsx",     xlsx_bytes)
    return zip_filename


# ─────────────────────────────────────────────────────────────────────────────
# 主程式
# ─────────────────────────────────────────────────────────────────────────────
def _print(msg: str) -> None:
    print(msg, flush=True)


def main() -> None:
    import sys

    if os.geteuid() != 0:
        _print("[警告] 建議以 root（sudo）執行，否則部分資料將無法收集。")
        _print("")

    computer_name = platform.node()
    local_ip      = get_local_ip_address()
    ts            = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    base_name     = f"{computer_name}_{local_ip}_{ts}"

    _print("=" * 56)
    _print("  H2C Linux 資安健診工具  v2.0")
    _print(f"  主機：{computer_name}   IP：{local_ip}")
    _print("=" * 56)
    _print("")

    steps = [
        ("[1/22]  取得系統資訊...",              get_system_info,          "系統資訊"),
        ("[2/22]  取得修補 / 更新狀態...",       get_update_status,        "更新狀態"),
        ("[3/22]  偵測防毒 / 端點防護...",       get_av_status,            "防毒工具"),
        ("[4/22]  取得已安裝套件...",            get_installed_packages,   "已安裝套件"),
        ("[5/22]  取得使用者帳號...",            get_local_user_accounts,  "使用者帳號"),
        ("[6/22]  取得特權帳號...",              get_privileged_accounts,  "特權帳號"),
        ("[7/22]  取得密碼 / PAM 原則...",       get_password_policy,      "密碼原則"),
        ("[8/22]  取得 SSH 安全設定...",         get_ssh_security,         "SSH安全設定"),
        ("[9/22]  取得網路設定...",              get_network_settings,     "網路設定"),
        ("[10/22] 取得啟動服務與排程...",        get_startup_items,        "啟動項目"),
        ("[11/22] 取得執行中程序...",            get_processes,            "執行中程序"),
        ("[12/22] 收集程序情資素材 / SHA256...", get_process_hashes,       "程序雜湊"),
        ("[13/22] 取得對外監聽 / 網路連線...",   get_netstat,              "網路連線"),
        ("[14/22] 取得防火牆狀態...",            get_firewall_status,      "防火牆狀態"),
        ("[15/22] 取得 Samba 設定...",           get_smb_status,           "SMB狀態"),
        ("[16/22] 取得網路分享 / 對外目錄...",   get_shared_folders,       "網路分享"),
        ("[17/22] 取得稽核與日誌設定...",        get_audit_policy,         "稽核原則"),
        ("[18/22] 取得 Kernel hardening...",     get_kernel_hardening,     "KernelHardening"),
        ("[19/22] 取得 SELinux / AppArmor...",   get_mac_status,           "MAC保護"),
        ("[20/22] 取得容器環境...",              get_container_status,     "容器環境"),
        ("[21/22] 取得檔案權限風險...",          get_file_permission_risks,"檔案權限風險"),
        ("[22/22] 讀取 hosts 檔案...",           get_hosts_file,           "Hosts檔案"),
    ]

    report_data: dict = {
        "meta": {
            "tool":          "H2C_PcSecCheck_linux",
            "version":       TOOL_VERSION,
            "schema_ver":    SCHEMA_VERSION,
            "computer_name": computer_name,
            "local_ip":      local_ip,
            "generated_at":  datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    }
    for label, func, key in steps:
        _print(label)
        try:
            report_data[key] = func()
        except Exception as e:
            _print(f"  [錯誤] {key}: {e}")
            report_data[key] = []

    _print("\n[分析] 執行風險評估...")
    findings               = analyze_findings(report_data)
    risk_score, risk_level = calculate_risk(findings)
    _print(f"[分析] 發現 {len(findings)} 個風險項目，等級: {risk_level.upper()} ({risk_score}/100)")

    _print("[打包] 產生 .h2cpc.zip...")
    zip_file = package_zip(base_name, report_data, findings, risk_score, risk_level, computer_name, local_ip)

    _print("")
    _print("=" * 56)
    _print(f"  完成！輸出：{zip_file}")
    _print(f"  風險等級：{risk_level.upper()} ({risk_score}/100)")
    for sev, label in [(4, "Critical"), (3, "High"), (2, "Medium"), (1, "Low"), (0, "Info")]:
        cnt = sum(1 for f in findings if f.get("severity") == sev)
        if cnt:
            _print(f"    {label:8s}: {cnt}")
    _print("=" * 56)
    _print("  請將 .h2cpc.zip 繳交給資安健診人員")
    _print("=" * 56)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[中斷] 使用者取消執行", flush=True)
    except Exception as e:
        print(f"\n[錯誤] {e}", flush=True)
        import traceback
        traceback.print_exc()
    input("\n按 Enter 鍵關閉視窗...")
